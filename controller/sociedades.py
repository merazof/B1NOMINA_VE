'''
Este archivo contiene las funciones básicas del CRUD de sociedades en el sistema
Created 2024-01
'''
'''
    **********************************************************************
    * Estructura del Modelo                                              *
    **********************************************************************
	id	bigint(20) AI PK
	rut	varchar(100)
	nombre	varchar(200)
	direccion	text
	region_id	bigint(20)
	comuna_id	bigint(20)
	ciudad	varchar(250)
	icono	varchar(250)
	created	datetime
	updated	datetime
	creator_user	bigint(20)
	updater_user	bigint(20)



    **********************************************************************
    * Estructura del Schema                                              *
    **********************************************************************
    rut : str = Field(min_length=3, max_length=100),
    nombre : str = Field(min_length=3, max_length=250),
    direccion : str = Field(min_length=3, max_length=500),
    region_id : int = Field(ge=1, le=1000),
    comuna_id : int = Field(ge=1, le=1000),
    ciudad : str = Field(min_length=3, max_length=250),
    icono : str = Field(min_length=3, max_length=250),

    model_config = {
        "json_schema_extra": {
            "examples": [
                {
                    "rut": "RutDemo",
                    "nombre":"Demo",
                    "direccion": "Direccion Demo",
                    "region_id": 1,
                    "comuna_id": 1,
                    "ciudad":"Demo ciudad",
                    'icono':''
                }
            ]
        }
    } 
'''   
import os
import random
import pdb

# import all you need from fastapi-pagination
from fastapi_pagination import Page, add_pagination
from fastapi_pagination.ext.sqlalchemy import paginate
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment


from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import select
from sqlalchemy import or_,and_


import  datetime


# importamos el modelo de la base de datos
from models.sociedades import Sociedad as SociedadModel
from models.historico_sociedades import HistoricoSociedad as HistoricoSociedadModel
from models.sede import Sede as SedeModel
from models.user import Usuario as UsuarioModel
from models.departamentos import Departamentos as DepartamentosModel
from models.grupos_empleados import GruposEmpleado as GruposEmpleadoModel
from models.categorias_configuracion import CategoriasConfiguracion as CategoriasConfiguracionModel
from models.configuracion import Configuraciones as ConfiguracionesModel
from models.view_general_user_sdg import ViewGeneralUserSDG  as ViewGeneralUserSDGModel
from models.view_general_user_sdg2 import ViewGeneralUserSDG2  as ViewGeneralUserSDG2Model
from models.view_general_user2 import ViewGeneralUser2  as ViewGeneralUser2Model
from models.view_contratados import ViewContratados as ViewContratadosModel
from models.view_no_contratados import ViewNoContratados as ViewNoContratadosModel
from models.cuentas_contable import CuentasContables as CuentasContablesModel
from models.tipos_prestamos import TiposPrestamos as TiposPrestamosModel
from models.bancos import Bancos as BancosModel
from models.estado_civil import EstadoCivil as EstadoCivilModel
from models.nacionalidad import Nacionalidad as NacionalidadModel
from models.cargo import Cargos as CargosModel
from models.regiones import Regiones as RegionesModel
from models.comunas import Comunas as ComunasModel
from models.nivel_estudio import NivelEstudio as NivelEstudioModel
from models.bancarios_sociedades import BancariosSociedad as BancariosSociedadModel
from models.historico_bancarios_sociedades import HistoricoBancariosSociedad as HistoricoBancariosSociedadModel
from models.mutuales_sociedades import MutualesSociedad as MutualesSociedadModel
from models.historico_mutuales_sociedades import HistoricoMutualesSociedad as HistoricoMutualesSociedadModel
from models.afp import AFP as AFPModel
from models.prevision_salud import PrevisionSalud as PrevisionSaludModel
from models.estatus_inscripcion import EstatusInscripcion as EstatusInscripcionModel
from models.grupos_centralizaciones import GrupoCentralizaciones as GrupoCentralizacionesModel
from models.campos_adicionales import CamposAdicionales as CamposAdicionalesModel


#importamos el esquema de datos de Sociedades
from schemas.sociedades import Sociedades as SocidadeSchema
from schemas.sociedades_basicos import SociedadesBasico as  SociedadesBasicoSchema
from schemas.sociedades_representante import SociedadesRepresentante as SociedadesRepresentanteSchema
from schemas.bancarios_sociedad import BancarioSociedad as BancarioSociedadSchema
from schemas.mutuales_sociedad import MutalesSociedad as MutalesSociedadSchema

from controller.bancarios_sociedad import BancariosSociedadController
from controller.mutuales_sociedad import MutualesSociedadController

# esto representa los metodos implementados en la tabla
class sociedadesController():
    # metodo constructor que requerira una instancia a la Base de Datos
    def __init__(self,db) -> None:
        self.db = db

    # funcion para crear el registro de historico las sociedades
    #@param sociedad: Modelo del registro de Sociedades
    #@param observavacion: Observación sobre el historico
    def create_historico_sociedad (self, sociedad: SociedadModel, observacion:str):
        # determinamos la fecha/hora actual
        ahora = datetime.datetime.now()

        try:
            #creamos la instancia la nuevo registro del historico
            newHistoricoSociedad= HistoricoSociedadModel(
                rut=sociedad.rut,
                sociedad_id=sociedad.id,
                nombre=sociedad.nombre,
                direccion=sociedad.direccion,
                region_id=sociedad.region_id,
                comuna_id=sociedad.comuna_id,
                ciudad=sociedad.ciudad,
                icono=sociedad.icono,
                email=(sociedad.email),
                responsable=(sociedad.responsable),
                rut_responsable=sociedad.rut_responsable,
                email_responsable=(sociedad.email_responsable),
                telefono_responsable=sociedad.telefono_responsable,
                created=sociedad.created,
                updated=sociedad.updated,
                creator_user=sociedad.creator_user,
                updater_user=sociedad.updater_user,
                fecha_registro = ahora,
                observaciones = observacion
            )

            # confirmamos el registro en el historico
            self.db.add(newHistoricoSociedad)
            self.db.commit()        

            result=True
            return (result)
        except ValueError as e:
            return( {"result":False,"error": str(e)})        
    
    
    #metodo para insertar  los datos de la sociedad
    # @userCreatorId: Id del usuario que está creando el registro
    # @params socieda: esquema de los datos de  sociedad que se desea insertar       
    def create_sociedad(self, sociedad:SocidadeSchema, userCreatorId:int ):
        #obtenemos la fecha/hora del servidor
        ahora=datetime.datetime.now()

        nombreSociedad=sociedad.nombre.upper().strip()
        rutSociedad=sociedad.rut.upper().strip()


        # contamos si existe una sociedad con el mismo nombre
        nRecordNombre = self.db.query(SociedadModel).filter(SociedadModel.nombre == nombreSociedad).count()  

        # contamos si existe una sociedad con el mismo rut
        nRecordRut = self.db.query(SociedadModel).filter(SociedadModel.rut == rutSociedad).count()         


        if (nRecordNombre > 0):
            # buscamos la sociedad con el nombre y lo devolvemos
            sociedadExists=self.db.query(SociedadModel).filter(SociedadModel.nombre == nombreSociedad).first() 

            # devolvemos la sociedad que ya existe
            return ({"result":"-1","estado":"Existe una sociedad con ese nombre","data":sociedadExists})
        elif (nRecordRut > 0):
            # buscamos la sociedad con el nombre y lo devolvemos
            sociedadExists=self.db.query(SociedadModel).filter(SociedadModel.rut == rutSociedad).first() 

            # devolvemos la sociedad que ya existe
            return ({"result":"-1","estado":"Existe una sociedad con ese rut","data":sociedadExists})            
        else:    
            #creamos el nuevo registro de banco
            try:
                newSociedad=SociedadModel(
                    nombre=((sociedad.nombre).upper()).strip(),
                    rut=((sociedad.rut).upper()).strip(),
                    region_id=sociedad.region_id,
                    comuna_id=sociedad.comuna_id,
                    direccion=((sociedad.direccion).upper()).strip(),
                    ciudad=((sociedad.ciudad).upper()).strip(),
                    icono=sociedad.icono,
                    email=(sociedad.email).strip().lower(),
                    responsable=(sociedad.responsable).strip().upper(),
                    rut_responsable=sociedad.rut_responsable,
                    email_responsable=(sociedad.email_responsable).strip().lower(),
                    telefono_responsable=sociedad.telefono_responsable,
                    created=ahora,
                    updated=ahora,
                    creator_user = userCreatorId,
                    updater_user=userCreatorId
                )

                #confirmamos el cambio en la Base de Datos
                self.db.add(newSociedad)
                self.db.commit()

                #creamos el registro historico de sede
                self.create_historico_sociedad(newSociedad,"Se creó una sociedad en el sistema")
                
                #llenamos el esquema de los datos bancarios de la sociedad
                bancariosSociedad=BancarioSociedadSchema(
                    sociedad_id=newSociedad.id,
                    banco_id=sociedad.banco_id,
                    numero_cuenta=sociedad.numero_cuenta,
                    codigo_convenio=sociedad.codigo_convenio,
                    giro_empresa=sociedad.giro_empresa,
                    ocultar_mail=sociedad.ocultar_email    
                )
                
                #insertamos el registro de datos bancarios de la sociedad
                resultaBancarios=BancariosSociedadController.create_bancario_sociedad(self,bancariosSociedad,userCreatorId)
                
                #llenamos el esquema de datos mutuales de la empresa
                mutualSociedad = MutalesSociedadSchema (
                    sociedad_id= newSociedad.id,
                    mutual_id=sociedad.mutual_id,
                    porcentaje=sociedad.porcentaje
                )
                
                #insertamos el registro de datos mutuales de la empresa
                resultMutual=MutualesSociedadController.create_mutual_sociedad(self,mutualSociedad,userCreatorId)
                
                
                #devolvemos un json con los datos insertados  
                data={
                    "id":newSociedad.id,
                    "rut":newSociedad.rut,
                    "nombre": newSociedad.nombre,
                    "region_id":newSociedad.region_id,
                    "comuna_id":newSociedad.comuna_id,
                    "ciudad":newSociedad.ciudad,
                    "direccion":newSociedad.direccion,
                    "icono":newSociedad.icono,
                    "email":newSociedad.email,
                    "responsable":newSociedad.responsable,
                    "rut_responsable":newSociedad.rut_responsable,
                    "email_responsable":newSociedad.email_responsable,
                    "telefono_responsable":newSociedad.telefono_responsable,
                    "banco_id": sociedad.banco_id,
                    "numero_cuenta":sociedad.numero_cuenta,
                    "codigo_convenio":sociedad.codigo_convenio,
                    "giro_empresa":sociedad.giro_empresa,
                    "ocultar_mail":sociedad.ocultar_email,   
                    "mutual_id":sociedad.mutual_id,
                    "porcentaje":sociedad.porcentaje,                                       
                    "created": newSociedad.created.strftime("%Y-%m-%d %H:%M:%S"),  
                    "updated":newSociedad.updated.strftime("%Y-%m-%d %H:%M:%S"),  
                    "creator_user":newSociedad.creator_user,
                    "updater_user":newSociedad.updater_user
                }

                return ({"result":"1","estado":"creado","data":data})
            except ValueError as e:
                return( {"result":"-3","error": str(e)})
    

    #metodo para consultar los datos de una sociedad por Id
    # @userUpdaterId: Id del usuario que está actualizando el registro
    def get_sociedad(self,id:int):

        # buscamos si esta sociedad existe
        nRecord = self.db.query(SociedadModel).filter(SociedadModel.id == id).count()
        
        if (nRecord == 0):
            # no existen datos de esta sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            # se extraen los datos de la sociedad
            try:
                sociedadExits = self.db.query(SociedadModel).filter(SociedadModel.id == id).first()                  
                
                
                #buscamos los datos bancarios de la sociedad
                resultBancarioSociedad=BancariosSociedadController.get_bancario_sociedad(self,id)
                if (resultBancarioSociedad['result']=="1"):
                    dataBancario=resultBancarioSociedad['data']
                else:
                    dataBancario={
                        "banco_id":"",
                        "numero_cuenta":"",
                        "codigo_convenio":"",
                        "giro_empresa":"",
                        "razon_social":"",
                        "ocultar_email":0
                    }    
                
                
                #buscamos los datos de los mutuales
                resultMutualSociedad=MutualesSociedadController.get_mutual_sociedad(self,id)
                
                if (resultMutualSociedad['result']=="1"):
                    data=resultMutualSociedad["data"]
                    dataMutual={
                        "mutual_id":data.mutual_id,
                        "porcentaje":data.porcentaje
                    }                                        
                else:
                    dataMutual={
                        "mutual_id":0,
                        "porcentaje":0
                    }
                
                
                #devolvemos un diccionario con los datos encontrados 

                data={
                    "id":sociedadExits.id,
                    "rut":sociedadExits.rut,
                    "nombre": sociedadExits.nombre,
                    "region_id":sociedadExits.region_id,
                    "comuna_id":sociedadExits.comuna_id,
                    "ciudad":sociedadExits.ciudad,
                    "direccion":sociedadExits.direccion,
                    "icono":sociedadExits.icono,
                    "email":sociedadExits.email,
                    "responsable":sociedadExits.responsable,
                    "rut_responsable":sociedadExits.rut_responsable,
                    "email_responsable":sociedadExits.email_responsable,
                    "telefono_responsable":sociedadExits.telefono_responsable,
                    "banco_id": dataBancario.banco_id,
                    "numero_cuenta":dataBancario.numero_cuenta,
                    "codigo_convenio":dataBancario.codigo_convenio,
                    "giro_empresa":dataBancario.giro_empresa,
                    "ocultar_mail":dataBancario.ocultar_email,   
                    "mutual_id":dataMutual.mutual_id,
                    "porcentaje":dataMutual.porcentaje,                                       
                    "created": dataBancario.created.strftime("%Y-%m-%d %H:%M:%S"),  
                    "updated":dataBancario.updated.strftime("%Y-%m-%d %H:%M:%S"),  
                    "creator_user":dataBancario.creator_user,
                    "updater_user":dataBancario.updater_user
                }                
                return ({"result":"1","estado":"Se consiguieron los datos de la sociedad","data":data})
            except ValueError as e:
                # ocurrió un error devolvemos el error
                return( {"result":"-1","error": str(e)}) 
            

    #metodo para consultar los datos de una sociedad por Id
    # @userUpdaterId: Id del usuario que está actualizando el registro
    def get_sociedad_basicos(self,id:int):

        # buscamos si esta sociedad existe
        nRecord = self.db.query(SociedadModel).filter(SociedadModel.id == id).count()
        
        if (nRecord == 0):
            # no existen datos de esta sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            # se extraen los datos de la sociedad
            try:
                sociedadExits = self.db.query(SociedadModel).filter(SociedadModel.id == id).first()                  
                
                #devolvemos un diccionario con los datos encontrados 

                data=sociedadExits.basicos_to_dict()
                           
                return ({"result":"1","estado":"Se consiguieron los datos de la sociedad","data":data})
            except ValueError as e:
                # ocurrió un error devolvemos el error
                return( {"result":"-1","error": str(e)})         


    #metodo para consultar los datos del representante sociedad por Id
    # @userUpdaterId: Id del usuario que está actualizando el registro
    def get_sociedad_representante(self,id:int):

        # buscamos si esta sociedad existe
        nRecord = self.db.query(SociedadModel).filter(SociedadModel.id == id).count()
        
        if (nRecord == 0):
            # no existen datos de esta sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            # se extraen los datos de la sociedad
            try:
                sociedadExits = self.db.query(SociedadModel).filter(SociedadModel.id == id).first()                  
                
                #devolvemos un diccionario con los datos encontrados 

                data=sociedadExits.representante_to_dict()
                           
                return ({"result":"1","estado":"Se consiguieron los datos de la sociedad","data":data})
            except ValueError as e:
                # ocurrió un error devolvemos el error
                return( {"result":"-1","error": str(e)})                   
            

    #metodo para efectuar búsquedas en las sociedades
    # @params cadena: cadena que se buscara en la tabla sociedades comparando con el campo nombre 
    def search_sociedades(self,finding ,page, records):

        findingT="%"+finding+"%"

        try:
            # buscamos si hay resgistros coincidentes
            nRecord=self.db.query(SociedadModel).filter(SociedadModel.nombre.like(findingT) | SociedadModel.rut.like(findingT)).count()  

            # hay registros
            if (nRecord >0):
                # ejecutamos la consulta
                consulta = self.db.query(SociedadModel).filter(SociedadModel.nombre.like(findingT) | SociedadModel.rut.like(findingT))
                consulta = consulta.limit(records)
                consulta = consulta.offset(records * (page - 1))
                result=consulta.all()
                
                # devolvemos los resultados
                return ({"result":"1","estado":"Se encontraron registros coincidentes con los criterios de búsqueda","data":result})
            else:
                # los filtros no arrojaron resultados
                 return ({"result":"-1","estado":"No record found"})            

        except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","error": str(e)})       
            
 
    #metodo para actualizar los datos de una sociedad por Id
    # @params userUpdaterId: Id del usuario que está actualizando el registro
    # @params sociedad: esquema de los datos de la sociedad 
    # @params id: Id de la sociedad que será actualizado
    def update_sociedad(self, sociedad:SocidadeSchema, userUpdaterId:int, id:int):
        #obtenemos la fecha/hora del servidor
        ahora=datetime.datetime.now()
        
        # buscamos si este sociedad existe
        nRecord = self.db.query(SociedadModel).filter(SociedadModel.id == id).count()
        
        if (nRecord == 0):
            # no se consiguieron los datos de la sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            try:
                #extraemos los datos para guardar el histórico
                sociedadExists = self.db.query(SociedadModel).filter(SociedadModel.id == id).first()             

                #creamos el registro historicos ociedades 
                self.create_historico_sociedad(sociedadExists ,"Actualización de la data de la sociedad")

                #registramnos los cambios en la tabla de sociedades
                sociedadExists.nombre=((sociedad.nombre).upper()).strip(),
                sociedadExists.rut=((sociedad.rut).upper()).strip(),
                sociedadExists.region_id=sociedad.region_id,
                sociedadExists.comuna_id=sociedad.comuna_id,
                sociedadExists.direccion=((sociedad.direccion).upper()).strip(),
                sociedadExists.ciudad=((sociedad.ciudad).upper()).strip(),
                sociedadExists.icono=sociedad.icono,
                sociedadExists.email=(sociedad.email).strip().lower(),
                sociedadExists.responsable=(sociedad.responsable).strip().upper(),
                sociedadExists.rut_responsable=sociedad.rut_responsable,
                sociedadExists.email_responsable=(sociedad.email_responsable).strip().lower(),
                sociedadExists.telefono_responsable=sociedad.telefono_responsable,                
                sociedadExists.updated=ahora,
                sociedadExists.updater_user=userUpdaterId               

                #confirmamos los cambios
                self.db.commit()

                data=sociedadExists.to_dict()
                # se actualizó el registro devolvemos el registro actualizado
                return ({"result":"1","estado":"Se actualizó el dato de de la sociedad","data":data})
            except ValueError as e:
                return( {"result":"-3","error": str(e)})                    


    # metodo para actualizar los datos de una sociedad por Id
    # @params userUpdaterId: Id del usuario que está actualizando el registro
    # @params sociedad: esquema de los datos de la sociedad 
    # @params id: Id de la sociedad que será actualizado
    def update_sociedad_basicos(self, sociedad:SociedadesBasicoSchema, userUpdaterId:int, id:int):
        #pdb.set_trace()
        #obtenemos la fecha/hora del servidor
        paso=1
        ahora=datetime.datetime.now()
        
        # buscamos si este sociedad existe
        paso=2
        nRecord = self.db.query(SociedadModel).filter(SociedadModel.id == id).count()
        
        if (nRecord == 0):
            # no se consiguieron los datos de la sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            try:
                paso=3
                #extraemos los datos para guardar el histórico
                sociedadExists = self.db.query(SociedadModel).filter(SociedadModel.id == id).first()             

                #creamos el registro historicos ociedades 
                paso=4
                self.create_historico_sociedad(sociedadExists ,"Actualización de la data de la sociedad")

                #registramnos los cambios en la tabla de sociedades
                paso=5
                sociedadExists.nombre=((sociedad.nombre).upper()).strip()
                sociedadExists.rut=((sociedad.rut).upper()).strip()
                sociedadExists.region_id=sociedad.region_id
                sociedadExists.comuna_id=sociedad.comuna_id
                sociedadExists.direccion=((sociedad.direccion).upper()).strip()
                sociedadExists.ciudad=((sociedad.ciudad).upper()).strip() 
                sociedadExists.email=sociedad.email              
                sociedadExists.updated=ahora
                sociedadExists.updater_user=userUpdaterId               

                #confirmamos los cambios
                paso=13
                self.db.commit()

                paso=14
                data=sociedadExists.basicos_to_dict()
                
                paso=15
                # se actualizó el registro devolvemos el registro actualizado
                return ({"result":"1","estado":"Se actualizó el dato de de la sociedad","data":data})
            except ValueError as e:
                return( {"result":"-3","cadenaError": f"Error en el sistema {str(e)} Paso:{paso}"})  


    # metodo para actualizar los datos de una sociedad por Id
    # @params userUpdaterId: Id del usuario que está actualizando el registro
    # @params sociedad: esquema de los datos de la sociedad 
    # @params id: Id de la sociedad que será actualizado
    def update_sociedad_representante(self, sociedad:SociedadesRepresentanteSchema, userUpdaterId:int, id:int):
        #pdb.set_trace()
        #obtenemos la fecha/hora del servidor
        paso=1
        ahora=datetime.datetime.now()
        
        # buscamos si este sociedad existe
        paso=2
        nRecord = self.db.query(SociedadModel).filter(SociedadModel.id == id).count()
        
        if (nRecord == 0):
            # no se consiguieron los datos de la sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            try:
                paso=3
                #extraemos los datos para guardar el histórico
                sociedadExists = self.db.query(SociedadModel).filter(SociedadModel.id == id).first()             

                #creamos el registro historicos ociedades 
                paso=4
                self.create_historico_sociedad(sociedadExists ,"Actualización de la data de la sociedad")

                #registramnos los cambios en la tabla de sociedades
                paso=5
                sociedadExists.responsable=((sociedad.responsable).upper()).strip()
                sociedadExists.rut_responsable=(sociedad.rut_responsable)
                sociedadExists.email_responsable=(sociedad.email_responsable)
                sociedadExists.telefono_responsable=(sociedad.telefono_responsable)
                sociedadExists.updated=ahora
                sociedadExists.updater_user=userUpdaterId               

                #confirmamos los cambios
                paso=13
                self.db.commit()

                paso=14
                data=sociedadExists.representante_to_dict()
                
                paso=15
                # se actualizó el registro devolvemos el registro actualizado
                return ({"result":"1","estado":"Se actualizó el dato de de la sociedad","data":data})
            except ValueError as e:
                return( {"result":"-3","cadenaError": f"Error en el sistema {str(e)} Paso:{paso}"})  
            


    # metodo para consultar todas las sociedades
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedades(self, page, records):
        consulta = self.db.query(SociedadModel)
        consulta = consulta.limit(records)
        consulta = consulta.offset(records * (page - 1))
        result=consulta.all()
        return (result)
    

    # metodo para consultar todas las sedes por sociedades
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_sedes(self, page, records,idSociedad):
        consulta = self.db.query(SedeModel).filter(SedeModel.sociedad_id==idSociedad)
        consulta = consulta.limit(records)
        consulta = consulta.offset(records * (page - 1))
        result=consulta.all()
        return (result)
    
    
    # metodo para consultar todas los grupos de unaa sociedad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_grupos_empleados(self, idSociedad):
        consulta = self.db.query(GruposEmpleadoModel).filter(GruposEmpleadoModel.sociedad_id==idSociedad)
        result=consulta.all()
        return (result)    
    

    # metodo para consultar todas las categorias de configiracion de una sociedad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_categorias_configuracion(self, idSociedad):
        result = self.db.query(CategoriasConfiguracionModel).filter(CategoriasConfiguracionModel.sociedad_id==idSociedad).order_by(CategoriasConfiguracionModel.id).all()
        return (result)  

    
    # metodo para consultar todas las categorias de configiracion de una sociedad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_configuraciones(self, idSociedad):
        result = self.db.query(ConfiguracionesModel).filter(ConfiguracionesModel.sociedad_id==idSociedad).order_by(ConfiguracionesModel.id).all()
        return (result)    
       

    # metodo para consultar todas los departamentos por Id de socidad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_departamentos(self, page, records,idSociedad):
        consulta = self.db.query(DepartamentosModel).filter(DepartamentosModel.sociedad_id==idSociedad)
        consulta = consulta.limit(records)
        consulta = consulta.offset(records * (page - 1))
        result=consulta.all()
        return (result)    


    # metodo para consultar todas los departamentos por Id de socidad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_cuentas_contables(self, idSociedad):
        consulta = self.db.query(CuentasContablesModel).filter(CuentasContablesModel.sociedad_id==idSociedad)
        result=consulta.all()
        return (result)    


    # metodo para consultar todas los empleados por Id de socidad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_empleados(self, idSociedad):
        consulta = self.db.query(ViewContratadosModel).filter(and_(ViewContratadosModel.sociedad_id == idSociedad,ViewContratadosModel.activo==1))
        result=consulta.all()

        data = [{
                "id": row.id,
                "nombre": row.nombres,
                "apellido_paterno": row.apellido_paterno,
                "apellido_materno": row.apellido_materno,
                "activo": row.activo,
                "cargo": row.cargo,
                "sueldo":row.sueldo,                
                "rut": row.rut
            } for row in result]

        return (data)   
    
    

    # metodo para consultar todas los registros de campos adicionales por sociedad
    # @params idSociedad: sociedad sobre la cual se consultaran los campos adicionales
    def list_campos_adicionales(self,idSociedad):
        try:
            paso=1
            nRecord = self.db.query(CamposAdicionalesModel).filter(CamposAdicionalesModel.sociedad_id==idSociedad).count()
            
            if (nRecord>0):
                paso=2
                result= self.db.query(CamposAdicionalesModel).filter(CamposAdicionalesModel.sociedad_id==idSociedad).first()
                
                paso=3
                data=[]
                '''
                id	bigint(20) AI PK
                sociedad_id	bigint(20)
                camuser1	varchar(200)
                activo1	int(11)
                camuser2	varchar(200)
                activo2	int(11)
                camuser3	varchar(200)
                activo3	int(11)
                camuser4	varchar(200)
                activo4	int(11)
                camuser5	varchar(200)
                activo5	int(11)
                created	datetime
                updated	datetime
                creator_user	bigint(20)
                updater_user	bigint(20)
                '''
                record={"nombre":result.camuser1,"activo":result.activo1}
                record2={"campo1":record}
                data.append(record2)
                
                record={"nombre":result.camuser2,"activo":result.activo2}
                record2={"campo2":record}
                data.append(record2)                

                record={"nombre":result.camuser3,"activo":result.activo3}
                record2={"campo3":record}
                data.append(record2)                

                record={"nombre":result.camuser4,"activo":result.activo4}
                record2={"campo4":record}
                data.append(record2)                

                record={"nombre":result.camuser5,"activo":result.activo5}
                record2={"campo5":record}
                data.append(record2)                

                
                return ({"result":"1","message": "Datos Encontrados","data":data})
            else:
                return ({"result":"-1","message": "No se han definido campos adicionales"})
        except ValueError as e:
            return( {"result":"-3","cadenaError": f"Ocurrió un error inesperadpo {str(e)} Paso:{paso}"}) 
    
    # metodo para consultar todas los no contratados por Id de socidad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_no_empleados(self, idSociedad):
        consulta = self.db.query(ViewNoContratadosModel).filter(ViewNoContratadosModel.sociedad_id == idSociedad)

        result=consulta.all()
        
        data=[]
        for row in result:
            # buscamos el avance de un inscrito
            idUser=row.id
            
            # verificamso si existe
            nRecord=self.db.query(EstatusInscripcionModel).filter(EstatusInscripcionModel.user_id==idUser).count()
            
            if (nRecord > 0):
                # si existe buscamos el registro
                result2=self.db.query(EstatusInscripcionModel).filter(EstatusInscripcionModel.user_id==idUser).first()
                #datos personales
                avance=0
                avance=int(result2.rut)+int(result2.nombre)+int(result2.apellido)+int(result2.nacionalidad)+int(result2.sexo)
                avance+=int(result2.fecha_nac)+int(result2.estado_civil)
                # datos de ubicacion
                avance+=int(result2.region)+int(result2.comuna)+int(result2.direccion)
                #datos de contacto
                avance+=int(result2.telefono)+int(result2.email)
                #datos de contrato
                avance+=int(result2.tipo_contrato)+int(result2.termino)+int(result2.fecha_contratacion)
                avance+=int(result2.salario_base)+int(result2.unidad_sueldo)+int(result2.monto_sueldo)
                avance+=int(result2.modalidad)+int(result2.dias_descanso)+int(result2.nivel_estudio)
                #datos de puesto de trabajo
                avance+=int(result2.sociedad)+int(result2.sede)+int(result2.departamento)+int(result2.cargo)+int(result2.grupo)
                #datos de pago
                avance+=int(result2.medio)+int(result2.banco)+int(result2.tipo_cuenta)+int(result2.numero)+int(result2.contrato)
                '''
                como calcular el avance
                sumamos las columnas de EstatusInscripcion
                de las cuales solo tomaremos en cuenta las siguientes
                (7) rut,nombre,apellido,nacionalidad,sexo,fecha_nac,estado_civil
                (3) region,comuna,direccion,
                (2) telefono,email,
                (6) tipo_contrato,termino,fecha_contratacion,salario_base,unidad_sueldo,monto_sueldo,
                (8) sociedad,sede,departamento,cargo,grupo,modalidad,dias_descanso,nivel_estudio,
                (6) medio,banco,tipo_cuenta,numero,foto,cv 
                -----
                32 
                avance = suma * 100 / 32          
                '''
                avanceUser=round(((avance * 100) / 32),0)
            else:
                avanceUser=0    
            
            item={
                "id": row.id,
                "nombre": row.nombres,
                "apellido_paterno": row.apellido_paterno,
                "apellido_materno": row.apellido_materno,
                "activo": row.activo,
                "cargo": row.cargo,
                "sueldo":row.sueldo,                
                "rut": row.rut,
                "cv_estatus":row.cv_estado,
                "contrato_estatus":row.contrato_estado,
                "avance":avanceUser                
            }
            data.append(item)
            
            '''        
            data = [{
                "id": row.id,
                "nombre": row.nombres,
                "apellido_paterno": row.apellido_paterno,
                "apellido_materno": row.apellido_materno,
                "activo": row.activo,
                "cargo": row.cargo,
                "sueldo":row.sueldo,                
                "rut": row.rut,
                "cv_estatus":row.cv_estado,
                "contrato_estatus":row.contrato_estado,
                "avance":random.randint(1, 100)
            } for row in result]
            '''

        return (data)      
    
    # metodo para consultar todas los no contratados por Id de socidad
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_sociedad_empleados_inactivos(self, idSociedad):
        consulta = self.db.query(ViewContratadosModel).filter(and_(ViewContratadosModel.sociedad_id == idSociedad,ViewContratadosModel.activo == 0))

        result=consulta.all()

        data = [{
                "id": row.id,
                "nombre": row.nombres,
                "apellido_paterno": row.apellido_paterno,
                "apellido_materno": row.apellido_materno,
                "activo": row.activo,
                "cargo": row.cargo,
                "sueldo":row.sueldo,                
                "rut": row.rut
            } for row in result]

        return (data)          
    

    # metodo para consultar todas las TiposPrestamoss
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def list_tipos_prestamos_sociedad(self,idSociedad):
        consulta = self.db.query(TiposPrestamosModel).filter(TiposPrestamosModel.sociedad_id==idSociedad)
        result=consulta.all()
        return (result)
        

    # metodo para ejecutar búsquedas en los usuarios usando una cadena,sede,departamento y grupo
    # @params finding: contenido json que se buscará entre los campos de la vista de usuarios
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def search_users_sdg(self,id,sede_id,departamento_id,grupo_id):
        '''
            Posibles campos de busqueda
            ----------------------------------------
            sede_id,
            departamento_id,
            grupo_id
        '''

        try:

            # solo se filtro por sede
            if ((sede_id!=0) and (departamento_id==0) and (grupo_id==0)):
                consulta=self.db.query(ViewGeneralUserSDGModel).filter(ViewGeneralUserSDGModel.sociedad_id==id).\
                filter(ViewGeneralUserSDGModel.sede_id==sede_id)

            # se aplico filtro por departamento
            if ((sede_id==0) and (departamento_id!=0) and (grupo_id==0)):   
                 consulta=self.db.query(ViewGeneralUserSDGModel).filter(ViewGeneralUserSDGModel.sociedad_id==id).\
                    filter(ViewGeneralUserSDGModel.departamento_id==departamento_id)
                 
            # se aplico filtro por grupo
            if ((sede_id==0) and (departamento_id==0) and (grupo_id!=0)):   
                 consulta=self.db.query(ViewGeneralUserSDGModel).filter(ViewGeneralUserSDGModel.sociedad_id==id).\
                    filter(ViewGeneralUserSDGModel.grupo_empleados_id==grupo_id)                 
                                 
            
            # se aplico filtro por sede y departamento
            if ((sede_id!=0) and (departamento_id!=0) and (grupo_id==0)):   
                 consulta=self.db.query(ViewGeneralUserSDGModel).filter(ViewGeneralUserSDGModel.sociedad_id==id).\
                    filter(ViewGeneralUserSDGModel.sede_id==sede_id).\
                    filter(ViewGeneralUserSDGModel.departamento_id==departamento_id)

            # se aplico filtro por sede y grupo
            if ((sede_id!=0) and (departamento_id!=0) and (grupo_id==0)):   
                 consulta=self.db.query(ViewGeneralUserSDGModel).filter(ViewGeneralUserSDGModel.sociedad_id==id).\
                    filter(ViewGeneralUserSDGModel.sede_id==sede_id).\
                    filter(ViewGeneralUserSDGModel.grupo_empleados_id==grupo_id)
                 
            # se aplico filtro por departamnto y grupo
            if ((sede_id==0) and (departamento_id!=0) and (grupo_id!=0)):   
                 consulta=self.db.query(ViewGeneralUserSDGModel).filter(ViewGeneralUserSDGModel.sociedad_id==id).\
                    filter(ViewGeneralUserSDGModel.departamento_id==departamento_id).\
                    filter(ViewGeneralUserSDGModel.grupo_empleados_id==grupo_id)                 

            # se aplicao filtro por sede, departamento y grupo
            if ((sede_id!=0) and (departamento_id!=0) and (grupo_id!=0)):   
                 consulta=self.db.query(ViewGeneralUserSDGModel).filter(ViewGeneralUserSDGModel.sociedad_id==id).\
                    filter(ViewGeneralUserSDGModel.sede_id==sede_id).\
                    filter(ViewGeneralUserSDGModel.departamento_id==departamento_id).\
                    filter(ViewGeneralUserSDGModel.grupo_empleados_id==grupo_id) 
                 
            # contamnos los registros
            nrecord=consulta.count()

            if (nrecord > 0):
                result=consulta.all()

                data = [{
                        "id": row.id,
                        "nombre": row.nombres,
                        "apellido_paterno": row.apellido_paterno,
                        "apellido_materno": row.apellido_materno,
                        "activo": row.activo,
                        "cargo": "Cargo",
                        "sueldo":str( random.randint(4500,  10000)),                
                        "rut": row.rut
                    } for row in result]

                   
                return ({"result":"1","estado":"Se encontraron registros coincidentes con los creiterios de búsqueda","data":data})
            else:
                # los filtros no arrojaron resultados
                 return ({"result":"-1","estado":"No record found"})
            
        except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","error": str(e)})
            

    # metodo para ejecutar búsquedas en los usuarios usando una cadena,sede,departamento y grupo
    # @params finding: contenido json que se buscará entre los campos de la vista de usuarios
    # @params page: pagina de los datos que se mostrará
    # @params records: cantidad de registros por página
    def search_users_sdg2(self,id,sede_id,departamento_id,grupo_id):
        '''
            Posibles campos de busqueda
            ----------------------------------------
            sede_id,
            departamento_id,
            grupo_id
        '''

        try:

            # solo se filtro por sede
            if ((sede_id!=0) and (departamento_id==0) and (grupo_id==0)):
                consulta=self.db.query(ViewGeneralUserSDG2Model).filter(ViewGeneralUserSDG2Model.sociedad_id==id).\
                filter(ViewGeneralUserSDG2Model.sede_id==sede_id)

            # se aplico filtro por departamento
            if ((sede_id==0) and (departamento_id!=0) and (grupo_id==0)):   
                 consulta=self.db.query(ViewGeneralUserSDG2Model).filter(ViewGeneralUserSDG2Model.sociedad_id==id).\
                    filter(ViewGeneralUserSDG2Model.departamento_id==departamento_id)
                 
            # se aplico filtro por grupo
            if ((sede_id==0) and (departamento_id==0) and (grupo_id!=0)):   
                 consulta=self.db.query(ViewGeneralUserSDG2Model).filter(ViewGeneralUserSDG2Model.sociedad_id==id).\
                    filter(ViewGeneralUserSDG2Model.grupo_empleados_id==grupo_id)                 
                                 
            
            # se aplico filtro por sede y departamento
            if ((sede_id!=0) and (departamento_id!=0) and (grupo_id==0)):   
                 consulta=self.db.query(ViewGeneralUserSDG2Model).filter(ViewGeneralUserSDG2Model.sociedad_id==id).\
                    filter(ViewGeneralUserSDG2Model.sede_id==sede_id).\
                    filter(ViewGeneralUserSDG2Model.departamento_id==departamento_id)

            # se aplico filtro por sede y grupo
            if ((sede_id!=0) and (departamento_id!=0) and (grupo_id==0)):   
                 consulta=self.db.query(ViewGeneralUserSDG2Model).filter(ViewGeneralUserSDG2Model.sociedad_id==id).\
                    filter(ViewGeneralUserSDG2Model.sede_id==sede_id).\
                    filter(ViewGeneralUserSDG2Model.grupo_empleados_id==grupo_id)
                 
            # se aplico filtro por departamnto y grupo
            if ((sede_id==0) and (departamento_id!=0) and (grupo_id!=0)):   
                 consulta=self.db.query(ViewGeneralUserSDG2Model).filter(ViewGeneralUserSDG2Model.sociedad_id==id).\
                    filter(ViewGeneralUserSDG2Model.departamento_id==departamento_id).\
                    filter(ViewGeneralUserSDG2Model.grupo_empleados_id==grupo_id)                 

            # se aplicao filtro por sede, departamento y grupo
            if ((sede_id!=0) and (departamento_id!=0) and (grupo_id!=0)):   
                 consulta=self.db.query(ViewGeneralUserSDG2Model).filter(ViewGeneralUserSDG2Model.sociedad_id==id).\
                    filter(ViewGeneralUserSDG2Model.sede_id==sede_id).\
                    filter(ViewGeneralUserSDG2Model.departamento_id==departamento_id).\
                    filter(ViewGeneralUserSDG2Model.grupo_empleados_id==grupo_id) 
                 
            # contamnos los registros
            nrecord=consulta.count()

            if (nrecord > 0):
                result=consulta.all()

                data = [{
                        "id": row.id,
                        "nombre": row.nombres,
                        "apellido_paterno": row.apellido_paterno,
                        "apellido_materno": row.apellido_materno,
                        "activo": row.activo,
                        "cargo": row.cargo,
                        "sueldo":row.sueldo,                
                        "rut": row.rut
                    } for row in result]

                   
                return ({"result":"1","estado":"Se encontraron registros coincidentes con los creiterios de búsqueda","data":data})
            else:
                # los filtros no arrojaron resultados
                 return ({"result":"-1","estado":"No record found"})
            
        except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","error": str(e)})            


    # metodo para listar los datos historicos  de una sociedad
    # @params id: Id de la sociedad que se esta consultando
    def list_history_sociedades(self, page:int, records: int, id:int):

        # buscamos si exite la sociendad
        nRecord = self.db.query(HistoricoSociedadModel).filter(HistoricoSociedadModel.sociedad_id == id).count()
        
        if (nRecord == 0):
            # el no se consiguieron datos historicos de la sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            # existen los datos historicos de la sociedad
            try:
                # ejecutamos la consulta
                consulta = self.db.query(HistoricoSociedadModel).filter(HistoricoSociedadModel.sociedad_id == id)
                consulta = consulta.limit(records)
                consulta = consulta.offset(records * (page - 1))
                listHistorySociedad=consulta.all()
               
                # se actualizó el registro devolvemos el registro encontrado
                return ({"result":"1","estado":"Se consiguieron los datos historicos de la sociedad ","data": listHistorySociedad})
            except ValueError as e:
                # ocurrió un error devolvemos el error
                return( {"result":"-1","error": str(e)})     


    # esta funcion permite contar los empleados, contratados e inactivos
    def get_employee_summary(self, idSociedad):
        nRecordUserActivos=self.db.query(ViewContratadosModel).filter(and_(ViewContratadosModel.sociedad_id==idSociedad,ViewContratadosModel.activo==1)).count()

        nRecordUserInactivos=self.db.query(ViewContratadosModel).filter(and_(ViewContratadosModel.sociedad_id==idSociedad,ViewContratadosModel.activo==0)).count()

        nRecordNoUserContratados=self.db.query(ViewNoContratadosModel).filter(ViewNoContratadosModel.sociedad_id==idSociedad).count()

        data={
            "empleados" : str(nRecordUserActivos),
            "contratacion" : str(nRecordNoUserContratados), 
            "inactivos" : str(nRecordUserInactivos)
        }

        return ({"result":"1","estado":"Datos encontrados","data":data})  


    # metodo para listar los datos historicos  de una sociedad
    # @params id: Id de la sociedad que se esta consultando
    def list_history_sociedades(self, page:int, records: int, id:int):

        # buscamos si exite la sociendad
        nRecord = self.db.query(HistoricoSociedadModel).filter(HistoricoSociedadModel.sociedad_id == id).count()
        
        if (nRecord == 0):
            # el no se consiguieron datos historicos de la sociedad
            return ({"result":"-2","estado":"No record found"})
        else:
            # existen los datos historicos de la sociedad
            try:
                # ejecutamos la consulta
                consulta = self.db.query(HistoricoSociedadModel).filter(HistoricoSociedadModel.sociedad_id == id)
                consulta = consulta.limit(records)
                consulta = consulta.offset(records * (page - 1))
                listHistorySociedad=consulta.all()
               
                # se actualizó el registro devolvemos el registro encontrado
                return ({"result":"1","estado":"Se consiguieron los datos historicos de la sociedad ","data": listHistorySociedad})
            except ValueError as e:
                # ocurrió un error devolvemos el error
                return( {"result":"-1","error": str(e)})     


    # esta funcion permite agrupar todos los parametros que serán necesarios en la creación del usuarios
    def get_parametros_crear_usuario(self, idSociedad):

        #consultamos todos los bancos
        try:
            # ejecutamos la consulta
            result = self.db.query(BancosModel).all()
            Bancos = [{
                    "id": row.id,
                    "nombre": row.nombre

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})             
        
        #consultamos todos las sedes por idSociedad
        try:
            # ejecutamos la consulta
            result = self.db.query(SedeModel).filter(SedeModel.sociedad_id==idSociedad).all()
            Sedes = [{
                    "id": row.id,
                    "nombre": row.nombre

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})             
        
        #consultamos todos los departamentos por idSociedad
        try:
            # ejecutamos la consulta
            result = self.db.query(DepartamentosModel).filter(DepartamentosModel.sociedad_id==idSociedad).all()
            Departamentos = [{
                    "id": row.id,
                    "nombre": row.nombre

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})   
        
          
        #consultamos todos los grupos por idSociedad
        try:
            # ejecutamos la consulta
            result = self.db.query(GruposEmpleadoModel).filter(GruposEmpleadoModel.sociedad_id==idSociedad).all()
            Grupos = [{
                    "id": row.id,
                    "nombre": row.nombre

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})             
              

        #consultamos todos los estado civiles
        try:
            # ejecutamos la consulta
            result = self.db.query(EstadoCivilModel).all()
            EstadoCivil = [{
                    "id": row.id,
                    "nombre": row.descripcion

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)}) 
        
        
        #consultamos las nacionalidades
        try:
            # ejecutamos la consulta
            result = self.db.query(NacionalidadModel).all()
            Nacionalidad = [{
                    "id": row.id,
                    "nombre": row.Nacionalidad

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})         
        
        
        #consultamos los cargos
        try:
            # ejecutamos la consulta
            result = self.db.query(CargosModel).filter(CargosModel.sociedad_id == idSociedad).all()
            Cargos = [{
                    "id": row.id,
                    "nombre": row.nombre

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)}) 
        
        #periodo salario
        tiposSalario=[{"id":"1","nombre":"Quincenal"},{"id":"2","nombre":"Mensual"},{"id":"3","nombre":"A término"}]         
        
        #Termino Contrato
        TerminoContrato=[{"id":"0","nombre":"Indefinido"},{"id":"1","nombre":"Definido"}]   
        
        #Tipo Contrato
        TipoContrato=[{"id":"1","nombre":"Empleado"},{"id":"2","nombre":"Pasante"}]           
        
        #buscamos las regiones
        try:
            # ejecutamos la consulta
            result = self.db.query(RegionesModel).all()
            Regiones = [{
                    "id": row.id,
                    "nombre": row.nombre

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})
                
        #buscamos las Comuna
        try:
            # ejecutamos la consulta
            result = self.db.query(ComunasModel).all()
            Localidad = [{
                    "id": row.id,
                    "nombre": row.nombre,
                    "idregion":row.region_id

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})
        
        #buscamos los Niveles de Estudio
        try:
            # ejecutamos la consulta
            result = self.db.query(NivelEstudioModel).all()
            NivelesEstudio = [{
                    "id": row.id,
                    "nombre": row.descripcion,
                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": str(e)})        

        unidadesSueldo=[
                {"id":"1","nombre":"$"},
                {"id":"2","nombre":"UF"}
                ]
        
        medioPago=[
            {"id":"1","nombre":"Transferencia"},
            {"id":"2","nombre":"Cheque"},
            {"id":"3","nombre":"Contado"}
             ]
        
        data={
            "bancos" : Bancos,
            "sede":Sedes,
            "departamentos":Departamentos,
            "grupos":Grupos,
            "estadocivil":EstadoCivil,
            "nacionalidad":Nacionalidad,
            "cargos":Cargos,
            "tiposalario":tiposSalario,
            "regiones":Regiones,
            "localidad":Localidad,
            "terminocontrato":TerminoContrato,
            "tipocontrato":TipoContrato,
            "nivelestudio":NivelesEstudio,
            "unidadessueldo":unidadesSueldo,
            "mediopago":medioPago
        }

        return ({"result":"1","estado":"Datos encontrados","parametros":data}) 
    
    
    # esta funcion permite agrupar todos los parametros que serán necesarios en la creación de sociedades
    def get_parametros_crear_sociedad(self):

        #pdb.set_trace()
        
        #buscamos las regiones
        paso=4
        try:
            # ejecutamos la consulta
            paso=5
            result = self.db.query(RegionesModel).all()
            paso=6
            Regiones = [{
                    "id": row.id,
                    "nombre": row.nombre

                } for row in result]                
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": f"Error {str(e)} Paso:{paso}" })
                
        #buscamos las Comuna
        paso=7
        try:
            paso=8
            # ejecutamos la consulta
            result = self.db.query(ComunasModel).all()
            paso=9
            Localidad = [{
                    "id": row.id,
                    "nombre": row.nombre,
                    "idregion":row.region_id

                } for row in result]    
                       
        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","error": f"Error {str(e)} Paso:{paso}" })
        
        paso=10
        data={
            "regiones":Regiones,
            "localidad":Localidad        }
        paso=11
        return ({"result":"1","estado":"Datos encontrados","parametros":data})     
    
    
    # esta funcion permite crear el archivo plantilla de carga de empleados masivo
    def dowload_file_bulk_user(self, idSociedad):
        #ruta = os.getenv("MASSIVE_USERS")+"/FormatoCargaEMP.xlsx"
        ruta="/var/www/html/massive_users/FormatoCargaEMP.xlsx"
        
        try:
            paso=1
            libro = Workbook()
            
            paso=2
            # ejecutamos la consulta de la regiones
            #consulta=self.db.execute("select nombre as Region,id as IdRegion from Regiones")
            consulta = self.db.query(RegionesModel).select_from(RegionesModel).with_entities(RegionesModel.nombre.label("Region"), RegionesModel.id.label("IdRegion"))
            
            paso=3
            # Obtener los resultados
            regionesResult = consulta.all()
            
            paso=4
            hoja_empleados = libro.active
            hoja_empleados.title = "CargaEmpleados"
            paso=5
            hoja_afp = libro.create_sheet("AFP")
            paso=6
            hoja_salud = libro.create_sheet("Prevision SALUD")
            paso=7
            hoja_comunas = libro.create_sheet("COMUNAS")
            paso=8
            hoja_regiones = libro.create_sheet("REGIONES")
            paso=9
            hoja_nacionalidad = libro.create_sheet("NACIONALIDADES")
            paso=10
            hoja_estudio = libro.create_sheet("Nivel de Estudios")        
            
            #rellenamos la hoja de regiones
            paso=11
            #encabezado
            encabezado=("Region","IdRegion")
            hoja_regiones.append(encabezado) 
            paso=12
            if (regionesResult is not None):
                for fila in regionesResult:
                    hoja_regiones.append(list(fila))            
               
            
            paso=13
            # ejecutamos la consulta de las AFP
            consulta = self.db.query(AFPModel).select_from(AFPModel).with_entities(AFPModel.nombre.label("Nombre"), AFPModel.id.label("Id_AFP"))
            
            paso=14
            # Obtener los resultados
            afpResult = consulta.all()
                        
            #rellenamos la hoja de AFP
            paso=15
            #encabezado
            encabezado=("Nombre","Id_AFP")
            hoja_afp.append(encabezado) 
            paso=16
            if (afpResult is not None):
                for fila in afpResult:
                    hoja_afp.append(list(fila))   
                    
            
            paso=17
            # ejecutamos la consulta de las Comunas
            consulta = self.db.query(ComunasModel).select_from(ComunasModel).with_entities(ComunasModel.nombre.label("Nombre"), ComunasModel.id.label("IdComuna"), ComunasModel.region_id.label("IdRegion"))
            
            paso=18
            # Obtener los resultados
            comunasResult = consulta.all()
                        
            #rellenamos la hoja de Comunas
            paso=19
            #encabezado
            encabezado=("Nombre","IdComuna","idRegion")
            hoja_comunas.append(encabezado) 
            paso=20
            if (comunasResult is not None):
                for fila in comunasResult:
                    hoja_comunas.append(list(fila))   
                    
                    
                    
            paso=21
            # ejecutamos la consulta de prevision Salud
            consulta = self.db.query(PrevisionSaludModel).select_from(PrevisionSaludModel).with_entities(PrevisionSaludModel.nombre.label("Nombre"), PrevisionSaludModel.id.label("Id_Prev_Salud"))
            
            paso=22
            # Obtener los resultados
            saludResult = consulta.all()
                        
            #rellenamos la hoja de Prevision Salud
            paso=23
            #encabezado
            encabezado=("Nombre","Id_Prev_Salud")
            hoja_salud.append(encabezado) 
            paso=24
            if (saludResult is not None):
                for fila in saludResult:
                    hoja_salud.append(list(fila))    
                    

            paso=25
            # ejecutamos la consulta de las nacionalidades
            consulta = self.db.query(NacionalidadModel).select_from(NacionalidadModel).with_entities(NacionalidadModel.Nacionalidad.label("Nacionalidad"), NacionalidadModel.id.label("IdNacionalidad"))
            
            paso=26
            # Obtener los resultados
            nacionalidadResult = consulta.all()
                        
            #rellenamos la hoja de Nivel de estudio
            paso=27
            #encabezado
            encabezado=("id","Nivel")
            hoja_nacionalidad.append(encabezado) 
            paso=28
            if (nacionalidadResult is not None):
                for fila in nacionalidadResult:
                    hoja_nacionalidad.append(list(fila))                                                     



            paso=29
            # ejecutamos la consulta de los niveles de estudio
            consulta = self.db.query(NivelEstudioModel).select_from(NivelEstudioModel).with_entities(NivelEstudioModel.id.label("idNivel"), NivelEstudioModel.descripcion.label("Nivel"))
            
            paso=30
            # Obtener los resultados
            estudioResult = consulta.all()
                        
            #rellenamos la hoja de Comunas
            paso=31
            #encabezado
            encabezado=("Nacionalidad","IdNacionalidad")
            hoja_estudio.append(encabezado) 
            paso=32
            if (estudioResult is not None):
                for fila in estudioResult:
                    hoja_estudio.append(list(fila))  


            #encabezado de la prima pagina
            encabezado=("Rut","FechaIncorporacion","Nombres","ApellidoPaterno","ApellidoMaterno","Region","Comuna","Direccion","Correo","FechaNacimiento","EstadoCivil","Sexo","Nacionalidad","TelFijo","Celular","SueldoBase","UnidadSueldoBase","Jubilado AFP","Id_AFP","Dejar en blanco","Dejar en Blanco","Afiliado_AFC","Id_Prev_Salud","Pactado_Salud","IdTipoContrato","AsignacionColacion","AsignacionMovilizacion","MedioDePago","IdGrupo","H_Ingreso","H_Salida","Ciudad","Tipo de Identificador (RUT)","Observaciones Generales","Campo Usr1","Campo Usr2","Campo Usr3","Campo Usr4","Campo Usr5","CorreoPersonal","Id Dim 1","Id Dim 2","Id Dim 3","Id Dim 4","Id Dim 5","IdNivelEstudios","Jefatura")             
            hoja_empleados.append(encabezado) 

            # se rellena de amarillo los campos requeridos
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")                      
            hoja_empleados["A1"].fill = fill
            hoja_empleados["B1"].fill = fill
            hoja_empleados["C1"].fill = fill
            hoja_empleados["D1"].fill = fill
            hoja_empleados["I1"].fill = fill
            hoja_empleados["AG1"].fill = fill
            hoja_empleados["AU1"].fill = fill

            #agregamos los comentarios necesarios a las celdas
            hoja_empleados["F1"].comment = Comment("Tomar un Valor del ID de la hoja Regiones", "Root")
            hoja_empleados["G1"].comment = Comment("Tomar un Valor del IdComuna de la hoja Comunas donde coincida con el Id de la Region","Root")
            hoja_empleados["I1"].comment = Comment("Campo Obligatorio","Root")            
            hoja_empleados["K1"].comment = Comment("1 Soltero\n2 Casado\n3 Viudo\n4Divorciado","Root")      
            hoja_empleados["L1"].comment = Comment("M = Masculino\nF = Femenino","Root") 
            hoja_empleados["M1"].comment = Comment("Tomar Id de Nacionalidad según hoja de Nacionalidades","Root") 
            hoja_empleados["Q1"].comment = Comment("$ = Pesos\nUF = Unidad de Fomento","Root")                      
            hoja_empleados["S1"].comment = Comment("Id según hoja de AFP","Root")      
            hoja_empleados["Q1"].comment = Comment("0 No Afiliado\n1 Afiliado","Root")   
            hoja_empleados["V1"].comment = Comment("0 No Afiliado\n1 Afiliado","Root")    
            hoja_empleados["AB1"].comment = Comment("1 = Contado\n2 = Transferencia\n3 = Cheque","Root")    
            hoja_empleados["Y1"].comment = Comment("1 Contrato Indefinido\n2 Plazo Fijo","Root")     
            hoja_empleados["AD1"].comment = Comment("Ejemplo 08:00","Root")
            hoja_empleados["AE1"].comment = Comment("Ejemplo 18:00","Root")          
            hoja_empleados["AF1"].comment = Comment("En palabras\nEjemplo Santiago","Root") 
            hoja_empleados["AU1"].comment = Comment("Colocar:\n0 = No es Jefatura\n1 = Es Jefatura","Root")                 
            
            #inmobilizamos la primera linea
            hoja_empleados.freeze_panes="A2"                                                                                           

            #guardamos el archivo            
            paso=100
            libro.save(ruta)
            return ({"result":"1","data":"massive_users/FormatoCargaEMP.xlsx"}) 
        
        except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","cadenaError": f"Error {str(e)} paso {paso}"})         
 
 
    
    # esta funcion permite extraer los grupos de centralizacion data completa
    def get_centralizaciones(self,idSociedad):
        paso=1
        try:
            nRecord=self.db.query(GrupoCentralizacionesModel).filter(GrupoCentralizacionesModel.sociedad_id==idSociedad).count()
            if (nRecord > 0):
                centralizacionesExists=self.db.query(GrupoCentralizacionesModel).filter(GrupoCentralizacionesModel.sociedad_id==idSociedad).all()
                data=centralizacionesExists
                return ({"result":"1","estado":"Datos encontrados","data":data})     
            else:
                return ({"result":"-1","estado":"o se han definido centralizaciones"})     

        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","cadenaError": f"Error {str(e)} Paso:{paso}" })            
        
    
    # esta funcion permite extraer los grupos de centralizacion data reducida
    def get_centralizaciones2(self,idSociedad):
        paso=1
        try:
            nRecord=self.db.query(GrupoCentralizacionesModel).filter(GrupoCentralizacionesModel.sociedad_id==idSociedad).count()
            if (nRecord > 0):
                centralizacionesExists=self.db.query(GrupoCentralizacionesModel).filter(GrupoCentralizacionesModel.sociedad_id==idSociedad).all()

                #creamos un a arreglo de listas con los datos que nos interesan 
                data = [{
                        "id": row.id,
                        "nombre": row.nombre
                    } for row in centralizacionesExists]

                return ({"result":"1","estado":"Datos encontrados","data":data})     
            else:
                return ({"result":"-1","estado":"o se han definido centralizaciones"})     

        except ValueError as e:
            # ocurrió un error devolvemos el error
            return( {"result":"-1","cadenaError": f"Error {str(e)} Paso:{paso}" })                    