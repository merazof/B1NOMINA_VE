'''
Este archivo contiene la rutina de carga masiva de usuarios
Created 2024-04
'''
 
import os
import re
import uuid
import io
import csv
import pdb


import openpyxl
from controller.validaciones_user import ValidationController
from controller.contact_users import contactUserController
from controller.ubication_users import ubicationUserController
from controller.datos_laborales import DatosLaboralesController
from controller.datos_pago import DatosPagoController
from controller.campos_adicionales_user import CamposUserController
from controller.validaciones_user import ValidationController
from controller.usuarios_afc import UsuariosAFCController
from controller.usuarios_afp import UsuariosAFPController
from controller.usuarios_prevision_salud import UsuariosPrevisionSaludController
from controller.colacion_usuarios import ColacionUsuariosController


# import all you need from fastapi-pagination
from fastapi_pagination import Page, add_pagination
from sqlalchemy import select
from fastapi_pagination.ext.sqlalchemy import paginate


from sqlalchemy import or_,and_
import datetime

#Importamos los modeloas necesarios
from models.user import Usuario as UsuarioModel
from models.historico_user import HistoricoUsuario as HistoricoUsuarioModel
from models.modulo import Modulo as ModuloModel
from models.datos_laborales import DatosLaborales as DatosLaboralesModel
from models.datos_pago import DatosPago as DatosPagoModel
from models.view_general_user import ViewGeneralUser
from models.view_general_user_modulos import viewGeneralUserModulo as viewGeneralUserModuloModel
from models.view_general_user3 import ViewGeneralUser3 as ViewGeneralUser3Model
from models.view_precarga_user import ViewPrecargaUser as ViewPrecargaUserModel
from models.contacto import Contacto as contactUserModel
from models.ubicacion import Ubicacion as UbicacionUserModel
from models.usuario_grupo import UsuariosGruposEmpleado as UsuariosGruposEmpleadoModel
from models.historico_usuario_grupo import HistoricoUsuariosGruposEmpleado as HistoricoUsuariosGruposEmpleadoModel
from models.usuario_sociedad import SociedadUsuario as SociedadUsuarioModel
from models.historico_usuario_sociedad import HistoricoSociedadUsuario as HistoricoSociedadUsuarioModel
from models.view_usuarios_grupos_empleados import viewUsuariosGruposEmpleados as viewUsuariosGruposEmpleadosModel
from models.pic_users import  PicUsuarios as PicUsuariosModel
from models.view_profile_user import ViewProfileUser as ViewProfileUserModel
from models.regiones import Regiones as RegionesModel
from models.comunas import Comunas as ComunasModel
from models.estatus_inscripcion import EstatusInscripcion as EstatusInscripcionModel


from schemas.user import User as UserSchema
from schemas.user2 import UserMassive as UserMassiveSchema
from schemas.contact_user2 import ContactUserMassive as contactUserSchema
from schemas.ubicacion_user2 import UbicacionUserMassive as ubicacionUserSchema
from schemas.datos_laborales import DatosLaborales as DatosLaboralesSchema
from schemas.datos_pago import DatosPago as DatosPagoSchema
from schemas.campos_adicionales_user import CamposAdicionalesUser as CamposAdicionalesUserSchema
from schemas.usuarios_grupo import UsuariosGruposEmpleado as UsuariosGruposEmpleadoSchema
from schemas.usuarios_afc import UsuariosAFC as UsuariosAFCSchema
from schemas.usuarios_afp import UsuariosAFP as UsuariosAFPSchema
from schemas.usuarios_prevision_salud import UsuariosPrevicionSalud as UsuariosPrevicionSaludSchema
from schemas.colacion_usuarios import ColacionUser as ColacionUserSchema
from schemas.usuarios_sociedad import UsuariosSociedad as UsuariosSociedadSchema
from schemas.estatus_inscripcion import EstatusInscripcion as EstatusInscripcionSchema


# importamos la utilidad para generar el hash del password
from utils.hasher import hash_password


# esto representa los metodos implementados en la tabla
class CargaMasivaController():
    # metodo constructor que requerira una instancia a la Base de Datos
    def __init__(self,db) -> None:
        self.db = db
        

   #metodo para guardar en el historico de usuarios sociedads
    #@params userUpdater: usuario que efectua la accion sobre el usuario
    #@params userSociedad: Registro de UsuaruioSociedades para guardar en el historico
    #@params observacion: Observacion que se guradará en el historico del usuario como refencia de la acción efectuada        
    def create_history_user_society (self, userSociedad:SociedadUsuarioModel, observacion: str):
        # determinamos la fecha/hora actual
        ahora = datetime.datetime.now()
        paso=1
        try:

            #creamos la instancia la nuevo registro del historico
            newHistoricoUserSociety=HistoricoSociedadUsuarioModel(
                sociedad_user_id=userSociedad.id,     
                sociedad_id=userSociedad.sociedad_id,
                user_id=userSociedad.user_id,
                created=userSociedad.created,
                updated=userSociedad.updated,
                creator_user = userSociedad.creator_user,
                updater_user=userSociedad.updater_user,
                fecha_registro=ahora,
                observaciones=observacion
            )

            # confirmamos el registro en el historico
            paso=2
            self.db.add(newHistoricoUserSociety)
            paso=3
            self.db.commit()        

            result=True
            return (result)
        except ValueError as e:
            return( {"result":False,"cadenaError": f"Error {str(e)} paso {paso}"})   
        
                
    # metodo que permite asignar un usuario a una sociedad aqui2
    def asignate_user_society(self,usuarioSociedad:UsuariosSociedadSchema, userCreatorId : int):
        ahora=datetime.datetime.now()
        
        paso=1
        userId=usuarioSociedad.user_id
        sociedadId=usuarioSociedad.sociedad_id
        
        paso=2
        #buscamos si el usuario no esta asociado es esta sociedad
        nRecord = self.db.query(SociedadUsuarioModel).filter(and_(SociedadUsuarioModel.user_id==userId,SociedadUsuarioModel.sociedad_id==sociedadId)).count()

        if (nRecord > 0):
            paso=3
            # existe no lo podemos volver a asignar
            return ({"result":"-1","estado":"Ya existe una asociacion entre el usuario y la sociedad, no la puede volver a crear","UserId":userId})
        else:
            try:
                paso=4
                # se puede crear
                '''
                `id` bigint(20) NOT NULL AUTO_INCREMENT,
                `sociedad_id` bigint(20) NOT NULL,  
                `user_id` bigint(20) DEFAULT NULL,
                `created` datetime NOT NULL,
                `updated` datetime NOT NULL,
                `creator_user` bigint(20) NOT NULL,
                `updater_user` bigint(20) NOT NULL,            
                '''
                newSociedadUsuario=SociedadUsuarioModel (
                    sociedad_id=sociedadId,
                    user_id=userId,
                    created=ahora,
                    updated=ahora,
                    creator_user= userCreatorId,
                    updater_user=userCreatorId
                )
                
                paso=5
                
                #confirmamos el registro
                self.db.add(newSociedadUsuario)
                self.db.commit()
                
                #creamos el registro historico
                paso=6
                result2=CargaMasivaController.create_history_user_society(self,newSociedadUsuario,"Creacion de un registro en Usuarios Sociedades en el sistema")
            
                return ({"result":"1","estado":"Se creo el registro Usuarios Sociedad","data":newSociedadUsuario})                
            
            except ValueError as e:
                return( {"result":"-3","cadenaError": f"Error {str(e)} paso {paso}"})            
        

   #metodo para guardar en el historico del usuario usuarios grupos
    #@params userUpdater: usuario que efectua la accion sobre el usuario
    #@params user: Registro de user para guardar en el historico
    #@params observacion: Observacion que se guradará en el historico del usuario como refencia de la acción efectuada        
    def create_history_user_group (self, usuarioGrupo:UsuariosGruposEmpleadoModel, observacion: str):
        # determinamos la fecha/hora actual
        ahora = datetime.datetime.now()
        paso=1
        try:

            #creamos la instancia la nuevo registro del historico
            newHistoricoUserGrupo=HistoricoUsuariosGruposEmpleadoModel(
                usuarios_grupo_empleados_id=usuarioGrupo.id,     
                sociedad_id=usuarioGrupo.sociedad_id,
                grupo_empleados_id=usuarioGrupo.grupo_empleados_id,
                user_id=usuarioGrupo.user_id,
                created=usuarioGrupo.created,
                updated=usuarioGrupo.updated,
                creator_user = usuarioGrupo.creator_user,
                updater_user=usuarioGrupo.updater_user,
                fecha_registro=ahora,
                observaciones=observacion
            )

            # confirmamos el registro en el historico
            paso=2
            self.db.add(newHistoricoUserGrupo)
            paso=3
            self.db.commit()        

            result=True
            return (result)
        except ValueError as e:
            return( {"result":False,"cadenaError": f"Error {str(e)} paso {paso}"}) 
        
                
    # metod que permite asignar un usuario a un grupo
    # @params userGroup : Esquema de datos que permitye registrar una asginacion de usuarios a un grupo
    # @params userUpdater: id del Usuario que efectua la  asignacion a un grupo            
    # @params idGroup: id del Grupo al cual se asi8gna el empleado            
    def asignate_user_group(self,userGroup: UsuariosGruposEmpleadoSchema, userCreator):
        # buscamos si existe una asginacion  previa
        paso=1
        userId=userGroup.user_id

        paso=2
        nRecordUserGroup=self.db.query(UsuariosGruposEmpleadoModel).filter(UsuariosGruposEmpleadoModel.user_id==userId).count()

        if (nRecordUserGroup > 0):
            # esta asignado a un grupo no puede volver a crearlo
            #devolvemos el grupo al cual esta asignado
            userGroupExists=self.db.query(UsuariosGruposEmpleadoModel).filter(UsuariosGruposEmpleadoModel.user_id==userId).first()
            return ({"result":"-2","estado":"Este empleado ya esta asignado a un grupo de empleados","data":userGroupExists})
        else:
            # puede crear el registro
            try:
                paso=3
                newUserGroup = UsuariosGruposEmpleadoModel(
                    sociedad_id=userGroup.sociedad_id,
                    user_id=userGroup.user_id,
                    grupo_empleados_id=userGroup.grupo_empleados_id,
                    created=datetime.datetime.now(),
                    updated=datetime.datetime.now(),
                    creator_user=userCreator,
                    updater_user=userCreator
                )

                paso=4
                self.db.add(newUserGroup)
                paso=5
                self.db.commit()  

                #creamos el historico del usuariso grupos 
                paso=6
                CargaMasivaController.create_history_user_group(newUserGroup,"Asignacion de un usuario a un grupo")  

                return( {"result":"1","estado":"Usuario asignado a un grupo"})  

            except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","cadenaError": f"Error {str(e)} paso {paso}"})           

    #metodo para guardar en el historico del usuario
    #@params userUpdater: usuario que efectua la accion sobre el usuario
    #@params user: Registro de user para guardar en el historico
    #@params observacion: Observacion que se guradará en el historico del usuario como refencia de la acción efectuada        
    def create_history_user (self, user:UsuarioModel, observacion: str):
        # determinamos la fecha/hora actual
        ahora = datetime.datetime.now()
        paso=1
        try:
            #creamos la instancia la nuevo registro del historico
            newHistoricoUser=HistoricoUsuarioModel(
                user_id=user.id,     
                rut=user.rut,
                rut_provisorio=user.rut_provisorio,
                nombres = ((user.nombres).upper()).strip(),
                apellido_paterno = ((user.apellido_paterno).upper()).strip(),
                apellido_materno = ((user.apellido_materno).upper()).strip(),
                fecha_nacimiento = user.fecha_nacimiento,
                sexo_id=user.sexo_id,
                estado_civil_id=user.estado_civil_id,
                nacionalidad_id=user.nacionalidad_id,
                username=user.username,
                password=user.password,
                activo=user.activo,
                created=user.created,
                updated=user.updated,
                creator_user = user.creator_user,
                updater_user=user.updater_user,
                fecha_registro=ahora,
                observaciones=observacion
            )

            # confirmamos el registro en el historico
            self.db.add(newHistoricoUser)
            self.db.commit()        

            result=True
            return (result)
        except ValueError as e:
            return( {"result":False,"cadenaError": f"Error {str(e)} paso {paso}"})     
        
   #metodo para insertar  los datos personales del usuario   
    # @params usuario: esquema de los datos del usuario que se desea insertar       
    def create_user(self, usuario:UserSchema, creatorUserId):
        #validaciones de los datos 
        rutOk=False
        rutProvisorioOk=False
        nombreOk=False
        apellidoPaternoOk=False
        apellidoMaternoOk=False

        #validamos el rut
        paso=0   
        rutOk=ValidationController.validarRut(usuario.rut)
        
        #validamos el rutprovisorio
        if (len(usuario.rut_provisorio.strip())>0):
            rutProvisorioOk=ValidationController.validarRut(usuario.rut_provisorio)
        else:
            rutProvisorioOk=True    
        
        #validamos el nombre
        paso=1
        nombreOk=ValidationController.validar_nombre(usuario.nombres.strip().upper())
        
        #validamos el apellido paterno
        paso=2
        apellidoPaternoOk=ValidationController.validar_nombre(usuario.apellido_paterno.strip().upper())
        
        #validamos el apellido materno
        paso=3
        if (len(usuario.apellido_materno.strip())>0):
            apellidoMaternoOk=ValidationController.validar_nombre(usuario.apellido_materno.strip().upper())
        else:
            apellidoMaternoOk=True    

        #validamos que los datos primarios esten validados
        paso=4
        if (rutOk and nombreOk and apellidoPaternoOk and apellidoMaternoOk and rutProvisorioOk):
            #obtenemos la fecha/hora del servidor
            ahora=datetime.datetime.now()
            
            # buscamos si el rut o el rut provisiorio ya existen
            rutV=usuario.rut
            rutProvisorio=usuario.rut_provisorio.strip()
            userName=usuario.username

            # contamos si existe un username identico en la base de datos
            paso=5
            nRecordUserName = self.db.query(UsuarioModel).filter(UsuarioModel.username == userName).count()  

            # contamos si existe un rut identico en la base de datos 
            paso=6
            nRecordUserRut = self.db.query(UsuarioModel).filter(UsuarioModel.rut == rutV).count()

            # inicializamos el arreglo userExistsUserRutProvisorio para determiniar si hay rut provisiior en la base de dtos
            userExistsUserRutProvisorio=[]
            
            # verificamos que se haya suministrado un rut provisorio
            nRecordUserRutProvisorio=0
            if (len(rutProvisorio)>0):
                # contamos si hay un rut provisiorio en la base de datos que coincida con el sumnistrado por el usuario
                paso=7
                nRecordUserRutProvisorio = self.db.query(UsuarioModel).filter(UsuarioModel.rut_provisorio == rutProvisorio).count() 
                

            if (nRecordUserName > 0):
                # el username esta ocupado
                paso=8
                userExistsUserName = self.db.query(UsuarioModel).filter(UsuarioModel.username == userName).first()  
                return ({"result":"-2","estado":"Username ya existe, no puede volver a crearlo","userId": userExistsUserName.id,"userName":userExistsUserName.username })


            if ((nRecordUserRut > 0 ) or (nRecordUserRutProvisorio > 0)):
                # el rut ya existe
                paso=9
                userExistsUserRut = self.db.query(UsuarioModel).filter(UsuarioModel.rut == rutV).first()
                return ({"result":"-3","estado":"Este Rut ya existe en la Base de Datos, no puede volver a crearlo","userId": userExistsUserRut.id,"rut":rutV })                
            

            # no existe el rut, procedemos a insertar el registro
            try:
                paso=10
                newUser=UsuarioModel(
                    rut=usuario.rut,
                    rut_provisorio=usuario.rut_provisorio,
                    nombres = ((usuario.nombres).upper()).strip(),
                    apellido_paterno = ((usuario.apellido_paterno).upper()).strip(),
                    apellido_materno = ((usuario.apellido_materno).upper()).strip(),
                    fecha_nacimiento = usuario.fecha_nacimiento,
                    sexo_id=usuario.sexo_id,
                    estado_civil_id=usuario.estado_civil_id,
                    nacionalidad_id=usuario.nacionalidad_id,
                    username=usuario.username,
                    password=hash_password(usuario.password),
                    activo=False,
                    created=ahora,
                    updated=ahora,
                    creator_user = creatorUserId,
                    updater_user=creatorUserId 
                )
                self.db.add(newUser)
                self.db.commit()

                #insertamos el registro en el historico
                paso=11
                CargaMasivaController.create_history_user (self,newUser, "Creación del Usuario por importación masiva")            

                newUserId=newUser.id
                return ({"result":"1","estado":"creado","newUserId":newUserId})
            except ValueError as e:
                return( {"result":"-1","cadenaError": f"Error {str(e)} paso {paso}"})
        
        else:
            cadenaError="Existen errores de formatro en los datos:"
            if (not rutOk):
                cadenaError=cadenaError + " El Rut está mal formateado"
            if (not rutProvisorioOk):
                cadenaError=cadenaError + " El Rut Provisorio está mal formateado"                    

            if (not nombreOk):
                cadenaError=cadenaError + " El Nombre tiene caracteres inválidos"
            
            if (not apellidoPaternoOk):
                cadenaError=cadenaError + " El apellido Paterno tiene caracteres inválidos"

            if (not apellidoMaternoOk):
                cadenaError=cadenaError + " El apellido Materno tiene caracteres inválidos"                    

            return( {"result":"-6","cadenaError":cadenaError}) 
        
                
        
    # metodo para subir datos masivos de usuarios al servidor
    # @params creatorUserId: usuario que subio el archivo
    # @params file: archivo que se está subiendo al archivo
    async def upload_massive_user(self,sociedadId,creatorUserId,file):
        paso=1
        #obtenemos la fecha/hora del servidor
        ahora=datetime.datetime.now()
        
        try:
            # declaramos la ruta de almacenaje de las fotos del usuario   
            paso=2         
            ruta = os.getenv("MASSIVE_USERS")

            #diccionario que contiene los tipos archivos permitidos
            permitedExtensionMassiveUsers=  os.getenv("PERMITED_MASSIVE_FILES_USERS") 
            
        except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","cadenaError": f"Error {str(e)} paso {paso}"})



        # Guarda el archivo en el directorio "MASSIVE_USERS"
        try:
            paso=3
            slug = os.path.splitext(file.filename)[0]
        except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","cadenaError": f"Error {str(e)} paso {paso}"})                
         

        # Reemplaza los caracteres no deseados por caracteres seguros
        safeFilename = re.sub(r"[^a-zA-Z0-9_-]", "_", slug)+str(uuid.uuid4())

        #path = os.path.join("files_users", file.filename)
        try:
            paso=4
            path = os.path.join(ruta, f"{safeFilename}.{file.filename.split('.')[-1]}")
        except ValueError as e:
                # ocurrio un error y devolvemos el estado
                return( {"result":"-3","cadenaError": f"Error {str(e)} paso {paso}"})         

        
        #devolvemos la extensión para verificr si se puede o no guardar el archivo
        fileExtension=file.filename.split('.')[-1]

        if (fileExtension in permitedExtensionMassiveUsers):
            try:
                paso=10
                file_bytes = await file.read()
                paso=5
                paso=11
                file_stream = io.BytesIO(file_bytes)


                paso=12
                wb = openpyxl.load_workbook(file_stream)
                sheet = wb['CargaEmpleados']


                #definimos el arreglo de resultados
                resultados=[]
                
                #determinar la cantida de filas en el objeto sheet
                paso=13
                nRecord=sheet.max_row
                if (nRecord > 1):
                    paso=14
                    for row in sheet.iter_rows(min_row=2,values_only=True):  # Omitir primera fila (encabezados)
                        paso=15
                        if (row is not None):
                            estatusError=""
                            estatusOk=""                            
                            paso=16
                            # Obtener valores de las celdas
                            # esta es la estructura del archivo excel
                            '''
                            0	Rut
                            1	FechaIncorporacion
                            2	Nombres
                            3	ApellidoPaterno
                            4	ApellidoMaterno
                            5	Region
                            6	Comuna
                            7	Direccion
                            8	Correo
                            9	FechaNacimiento
                            10	EstadoCivil
                            11	Sexo
                            12	Nacionalidad
                            13	TelFijo
                            14	Celular
                            15	SueldoBase
                            16	UnidadSueldoBase
                            17	Jubilado AFP
                            18	Id_AFP
                            19	Dejar en blanco
                            20	Dejar en Blanco
                            21	Afiliado_AFC
                            22	Id_Prev_Salud
                            23	Pactado_Salud
                            24	IdTipoContrato
                            25	AsignacionColacion
                            26	AsignacionMovilizacion
                            27	MedioDePago
                            28	IdGrupo
                            29	H_Ingreso
                            30	H_Salida
                            31	Ciudad
                            32	Tipo de Identificador (RUT)
                            33	Observaciones Generales
                            34	Campo Usr1
                            35	Campo Usr2
                            36	Campo Usr3
                            37	Campo Usr4
                            38	Campo Usr5
                            39	CorreoPersonal
                            40	Id Dim 1
                            41	Id Dim 2
                            42	Id Dim 3
                            43	Id Dim 4
                            44	Id Dim 5
                            45	IdNivelEstudios
                            '''
                            if ((row[0] is not None) and (len(row[0].strip())>0)):  
                                paso=17
                                rutv=row[0]
                                paso=18
                                fechaInicio=row[1]
                                paso=19
                                if (row[2] is not None):
                                    paso=20
                                    nombresV = row[2].upper()
                                if (row[3] is not None):    
                                    paso=21
                                    apellidoPaterno = row[3].upper()
                                if (row[4] is not None):    
                                    paso=22
                                    apellidoMaterno = row[4].upper()

                                paso=23     
                                if (row[5] is not None)                               :
                                    region=row[5]
                                else:
                                    region=-1
                                    
                                paso=24
                                if (row[6] is not None):
                                    comuna=row[6]
                                else:
                                    comuna=-1
                                paso=25
                                if (row[7] is not None):
                                    direccion=row[7]
                                else:
                                    direccion=""
                                    
                                paso=26
                                if (row[8] is not None):  
                                    paso=27                
                                    correo = row[8].lower()
                                else:
                                    correo=""    
                                    
                                paso=28         
                                if (row[9] is not None):
                                    fechaNacimiento=row[9]
                                else:
                                    fechaNacimiento="1990-01-01"
                                    
                                paso=29
                                estadoCivil=row[10]
                                paso=30
                                sexo=row[11]
                                paso=31
                                nacionalidad=row[12]
                                paso=32
                                telefonoFijo=row[13]
                                paso=33
                                celular=row[14]
                                paso=34
                                sueldoBase=row[15]
                                paso=35
                                UnidadSueldoBase=row[16]
                                paso=36
                                JubiladoAFP=row[17]
                                paso=37
                                idAFP=row[18]
                                paso=38
                                Afiliado_AFC=row[21]
                                paso=39
                                Id_Prev_Salud=row[22]
                                paso=40
                                Pactado_Salud=row[23]
                                paso=41
                                IdTipoContrato=row[24]
                                paso=42
                                if (row[25] is not None):
                                    AsignacionColacion=row[25]
                                else:                                    
                                    AsignacionColacion=0
                                paso=43
                                if (row[26] is not None):
                                    AsignacionMovilizacion=row[26]
                                else:
                                    AsignacionMovilizacion=0
                                paso=44
                                if (row[27] is not None):
                                    MedioDePago=row[27]
                                else:
                                    MedioDePago="1" #predeterminamos efectivo
                                    
                                paso=45
                                IdGrupo=row[28]
                                paso=46
                                H_Ingreso=row[29]
                                paso=47
                                H_Salida=row[30]
                                paso=48
                                Ciudad=row[31]
                                paso=49
                                TipoIdentificador=row[32]
                                paso=50
                                Observaciones_Generales=row[33]
                                paso=51
                                Campo_Usr1=row[34]
                                paso=52
                                Campo_Usr2=row[35]
                                paso=53
                                Campo_Usr3=row[36]
                                paso=54
                                Campo_Usr4=row[37]
                                paso=55
                                Campo_Usr5=row[38]
                                paso=56
                                CorreoPersonal=row[39]
                                paso=57
                                IdDim1=row[40]
                                paso=58
                                IdDim2=row[41]
                                paso=59
                                IdDim3=row[42]
                                paso=60
                                IdDim4=row[43]
                                paso=61
                                IdDim5=row[44]
                                paso=62
                                IdNivelEstudios=row[45]
                                paso=63
                                esJefatura=row[46]

                                paso=64
                                if (IdNivelEstudios== None):
                                    IdNivelEstudios=6

                                allOk=True
                                causa=""


                                #validaciones
                                #formato del rut
                                paso=65
                                if (not ValidationController.validarRut(rutv)):
                                    allOk=False
                                    causa="Formato de Rut invalido"                    


                                #validar email
                                paso=66
                                if (not ValidationController.validarEmail(correo)):
                                    allOk=False
                                    causa="Correo invalido"       

                                
                                #validamos el nombre personal
                                paso=67
                                if (not ValidationController.validar_nombre2(nombresV)):
                                    allOk=False
                                    causa="El nombre personal posee caracteres inválidos"       

                                
                                #validamos el apellio paterno
                                paso=68
                                if (not ValidationController.validar_nombre2(apellidoPaterno)):
                                    allOk=False
                                    causa="El apellido paterno posee caracteres inválidos"    

                                
                                #validamos el apellio materno
                                paso=69
                                if (not ValidationController.validar_nombre2(apellidoMaterno)):
                                    allOk=False
                                    causa="El apellido materno posee caracteres inválidos"                     
                                                                
                                paso=70
                                if (sexo =="M"):
                                    sexoT=1
                                else:
                                    sexoT=2
                                    
                                # creamos el username sumando el primer nombre + "." + primer apellido
                                paso=71
                                mi_string = nombresV
                                if mi_string.isspace():
                                    arreglo = mi_string.split(' ')
                                    primer_elemento = arreglo[0].lower()
                                else:
                                    primer_elemento=mi_string.lower()

                                paso=72
                                mi_string = apellidoPaterno
                                if mi_string.isspace():
                                    paso=73
                                    arreglo = mi_string.split(' ')
                                    paso=74
                                    segundo_elemento = arreglo[0].lower()
                                else:
                                    paso=75
                                    segundo_elemento=mi_string.lower()

                                paso=76
                                # usernameV = (".".join([primer_elemento, segundo_elemento]))
                                usernameV=rutv

                                if (allOk): 
                                    paso=77                   
                                    Persona=UserMassiveSchema(rut=rutv,
                                                            rut_provisorio="",
                                                            nombres=nombresV,
                                                            apellido_paterno=apellidoPaterno,
                                                            apellido_materno=apellidoMaterno,
                                                            #fecha_nacimiento=datetime.datetime.strptime(fechaNacimiento, "%Y-%m-%d"),
                                                            fecha_nacimiento=fechaNacimiento,
                                                            sexo_id=sexoT,
                                                            estado_civil_id=estadoCivil,
                                                            nacionalidad_id=nacionalidad,
                                                            username=usernameV,
                                                            password=hash_password(rutv),
                                                            activo=True
                                                            )
                                    
                                    paso=78
                                    result=CargaMasivaController.create_user(self,Persona,creatorUserId)

                                    paso=79
                                    if (result['result']=="1"):
                                        # se inserto correctamente los datos personales
                                        newUserId=result["newUserId"]                                    
                                        
                                        #creamos el registro de estatus de inscripcion
                                        newEstatusInscripcion=EstatusInscripcionModel(
                                                user_id = newUserId,
                                                rut=1,
                                                nombre=1,    
                                                apellido=1,   
                                                nacionalidad=1,    
                                                sexo=1,   
                                                fecha_nac=1,    
                                                estado_civil=1,   
                                                # --ubicacion ----------------
                                                region=0,   
                                                comuna=0,    
                                                direccion=0,   
                                                # --contacto ------------------------------------
                                                telefono=0,    
                                                email=0,    
                                                # --laborales ---------------------------------# --
                                                tipo_contrato=0,   
                                                termino=0,    
                                                fecha_contratacion=0,   
                                                salario_base=0,    
                                                unidad_sueldo=0,   
                                                monto_sueldo=0,    
                                                sociedad=0,       
                                                sede=0,   
                                                departamento=0,   
                                                cargo=0,    
                                                grupo=0,   
                                                modalidad=0,    
                                                dias_descanso=0,
                                                nivel_estudio=0,    
                                                # --datos de pago -----------------------------------------# --
                                                medio=0,   
                                                banco=0,    
                                                tipo_cuenta=0,   
                                                numero=0,    
                                                # --archivos necesarios --------------------------------------------# --   
                                                foto=0,   
                                                cv=0,    
                                                contrato=0, 
                                                created = ahora,    
                                                updated = ahora,
                                                creator_user= creatorUserId,     
                                                updater_user = creatorUserId                                            
                                        )
                                        #confirmamos el registro
                                        paso=2000
                                        self.db.add(newEstatusInscripcion)
                                        paso=2001
                                        self.db.commit()                                         
                                        
                                        #determinamos el id del estatus de inscripcion para posteriores actualizaciones
                                        estatusInscripcionId=newEstatusInscripcion.id
                                        
                                        
                                        # insertamos datos de contacto
                                        paso=80
                                        contactoUsuario=contactUserSchema (
                                            user_id=newUserId,
                                            email=correo,
                                            fijo=telefonoFijo,
                                            movil=celular
                                        )
                                        paso=81
                                        result2=contactUserController.create_contact_user(self,creatorUserId ,contactoUsuario)
                                        #confirmamos el estatus de contacto en la tabla de estatus de inscripcion
                                        if (result2['result']=="1"):
                                            if (correo !=""):
                                                newEstatusInscripcion.email=1
                                            if ((telefonoFijo!="") or (celular!='')):
                                                newEstatusInscripcion.telefono=1
                                            self.db.commit()
                                            

                                        # insertamos datos de localizacion
                                        '''
                                        user_id: int = Field (ge=1, lt=20000)
                                        region_id: int = Field (ge=1, lt=20000)
                                        comuna_id: int = Field (ge=1, lt=20000)
                                        direccion: str = Field (min_length= 0, max_length= 250)                        
                                        '''
                                        paso=82
                                        if (region!=-1 and comuna!=-1):
                                            # comprobamos la region
                                            nRecordRegion=self.db.query(RegionesModel).filter(RegionesModel.id==region).count()
                                                                                       
                                            #comprobamos la comuna
                                            nRecordComuna=self.db.query(ComunasModel).filter(ComunasModel.id==region).count()
                                            
                                            if (nRecordRegion > 0 and nRecordComuna > 0):
                                                ubicacionUsuario=ubicacionUserSchema(
                                                    user_id=newUserId,
                                                    region_id=region,
                                                    comuna_id=comuna,
                                                    direccion=direccion
                                                )
                                                paso=83
                                                result3=ubicationUserController.create_ubication_user(self,creatorUserId,ubicacionUsuario)
                                                
                                                # actualizamos los datos de localizacion en el estatus de inscripcion
                                                newEstatusInscripcion.region=1
                                                newEstatusInscripcion.comuna=1
                                                if (direccion!=""):
                                                    newEstatusInscripcion.direccion=1
                                                
                                                self.db.commit()
                                                
                                            else:
                                                if ((nRecordRegion < 1) and (nRecordComuna <1)):
                                                    result3={'result':'-5'}    # la region y la comuna estan erradas
                                                elif ((nRecordRegion >0) and (nRecordComuna <1)):
                                                    result3={'result':'-6'}    # la region esta correcta y la comuna estan erradas
                                                elif ((nRecordRegion < 1) and (nRecordComuna >0)):
                                                    result3={'result':'-7'}    # la region esta equivocada y la comuna esta correcta
                                                else:
                                                    result3={'result':'-8'}  # Error inexperado
                                        
                                        #insertamos datos de laborales
                                        paso=84
                                        if (UnidadSueldoBase=="$"):
                                            unidadSueldo=1
                                        else:
                                            unidadSueldo=2

                                        '''
                                        IdGrupo=row[28]
                                        '''
                                        paso=85
                                        if (fechaInicio is not None) \
                                            and (IdTipoContrato is not None) \
                                            and (IdGrupo is not None)  \
                                            and (sueldoBase is not None)  \
                                            and (unidadSueldo is not None)  \
                                            and (IdNivelEstudios is not None) \
                                            and (H_Ingreso is not None) \
                                            and (H_Salida is not None):
                                            datosLaborales=DatosLaboralesSchema(
                                                sociedad_id=sociedadId,
                                                sede_id=1,
                                                departamento_id=1,
                                                grupo_id=IdGrupo,
                                                cargo_id=1,
                                                user_id=newUserId,
                                                tipo_contrato=IdTipoContrato,
                                                termino_contrato=1,
                                                #fecha_inicio= datetime.datetime.strptime(fechaInicio, "%Y-%m-%d").date(),
                                                fecha_inicio= fechaInicio,
                                                fecha_fin=datetime.datetime.strptime("1900-01-01", "%Y-%m-%d").date(),
                                                periodo_salario=30,
                                                modalidad=1,
                                                dias_descanso="6,7",
                                                salario_base=sueldoBase,
                                                nivel_estudio_id=IdNivelEstudios,
                                                unidad_sueldo=unidadSueldo,
                                                hora_ingreso= datetime.datetime.strptime(H_Ingreso, "%H:%M").time(),
                                                hora_egreso=datetime.datetime.strptime(H_Salida, "%H:%M").time(),
                                                jefatura=esJefatura
                                            )
                                            paso=35
                                            result4=DatosLaboralesController.create_datos_laborales(self,datosLaborales,creatorUserId)
                                            
                                            #actualizamos el estatus de inscripcion con los datos laborales
                                            newEstatusInscripcion.tipo_contrato=1
                                            newEstatusInscripcion.termino=1      
                                            newEstatusInscripcion.fecha_contratacion=1
                                            newEstatusInscripcion.salario_base=1
                                            newEstatusInscripcion.unidad_sueldo=1
                                            newEstatusInscripcion.monto_sueldo=1
                                            newEstatusInscripcion.sociedad=1
                                            newEstatusInscripcion.sede=1
                                            newEstatusInscripcion.departamento=1
                                            newEstatusInscripcion.cargo=1
                                            newEstatusInscripcion.grupo=1
                                            newEstatusInscripcion.modalidad=1
                                            newEstatusInscripcion.dias_descanso=1                                                                                                                              
                                            newEstatusInscripcion.nivel_estudio=1
                                            self.db.commit()

                                        else:
                                            result4={"result":"-100"} #no procesado
                                            
                                            
                                        #insertamos los datos de pago
                                        '''
                                        user_id	bigint(20)
                                        medio	int(11)
                                        banco_id	bigint(20)
                                        tipo_cuenta	int(11)                                
                                        '''
                                        paso=86
                                        datosPago=DatosPagoSchema(
                                            user_id=newUserId,
                                            medio=MedioDePago,
                                            banco_id=0,
                                            tipo_cuenta=0,
                                            numero_cuenta =''
                                        )
                                        paso=87
                                        result5=DatosPagoController.create_datos_pago(self,datosPago,creatorUserId)
                                        
                                        #actualizamos el estatus de inscripcion con los datos de pago
                                        newEstatusInscripcion.medio=1
                                        newEstatusInscripcion.banco=0
                                        newEstatusInscripcion.tipo_cuenta=0
                                        newEstatusInscripcion.numero=0
                                        self.db.commit()

                                        
                                        #insertamos los datos en los campos adicionales
                                        paso=88
                                        if ((Campo_Usr1 is not None) or (Campo_Usr2 is not None) or (Campo_Usr3 is not None) or (Campo_Usr4 is not None) or (Campo_Usr5 is not None) ):
                                            if (Campo_Usr1 is None):
                                                Campo_Usr1=""

                                            if (Campo_Usr2 is None):
                                                Campo_Usr2=""

                                            if (Campo_Usr3 is None):
                                                Campo_Usr3=""

                                            if (Campo_Usr4 is None):
                                                Campo_Usr4=""

                                            if (Campo_Usr5 is None):
                                                Campo_Usr5=""
                                            paso=89
                                            camposAdicionales=CamposAdicionalesUserSchema(
                                                sociedad_id=sociedadId,
                                                user_id=newUserId,
                                                camuser1=Campo_Usr1,
                                                camuser2=Campo_Usr2,
                                                camuser3=Campo_Usr3,
                                                camuser4=Campo_Usr4,
                                                camuser5=Campo_Usr5
                                            )
                                            result6=CamposUserController.create_campos_adicionales_user(self,camposAdicionales,creatorUserId)
                                        else:
                                            result6={"result":"-100"}
                                            
                                        
                                        # asignar a grupo
                                        '''
                                        IdGrupo=row[28]
                                        '''
                                        paso=90
                                        if (IdGrupo is not None ):    
                                            usuarioGrupo=UsuariosGruposEmpleadoSchema(
                                                sociedad_id=sociedadId,
                                                user_id=newUserId,
                                                grupo_empleados_id=IdGrupo
                                            )
                                            paso=91
                                            result7=CargaMasivaController.asignate_user_group(self,usuarioGrupo,creatorUserId)
                                        else:
                                            result7={"result":"-100"}
                                            #actualizamos el estatus de inscripcion con los datos de grupo

                                        #datos AFC
                                        '''
                                        Afiliado_AFC=row[21]
                                        '''
                                        paso=92
                                        afiliadoAFC=UsuariosAFCSchema(
                                            user_id=newUserId,
                                            estado=Afiliado_AFC,
                                            antiguedad=0,
                                            vejez=0,
                                            invalidez=0,
                                            exinp=0
                                        )
                                        paso=93
                                        result8=UsuariosAFCController.create_usuario_afc(self,creatorUserId,afiliadoAFC)
                                        
                                        #datos AFP
                                        '''
                                        JubiladoAFP=row[17]
                                        idAFP=row[18]                               
                                        '''
                                        paso=94
                                        if ((JubiladoAFP is not None) and (idAFP is not None)):
                                            #cargamos el esquema de datos
                                            AfiliadoAFP=UsuariosAFPSchema(
                                                user_id=newUserId,
                                                afp_id=idAFP,
                                                jubilado_afp=JubiladoAFP,
                                                ahorro_afp2=0
                                            )
                                            paso=95
                                            result9=UsuariosAFPController.create_usuario_afp(self,creatorUserId,AfiliadoAFP)
                                        else:    
                                            result9={"result":"-100"}

                                        
                                        #datos Prevision Salud
                                        '''
                                        Id_Prev_Salud=row[22]
                                        Pactado_Salud=row[23]
                                        IdTipoContrato=row[24]                                
                                        '''
                                        paso=96
                                        if (Id_Prev_Salud is not None):
                                            usuarioPrevision=UsuariosPrevicionSaludSchema(
                                                user_id=newUserId ,
                                                pactado=Pactado_Salud,
                                                prevision_salud_id= Id_Prev_Salud,
                                                tipo_contrato= IdTipoContrato
                                            )
                                            paso=97
                                            result10=UsuariosPrevisionSaludController.create_usuario_prevision_salud(self,creatorUserId,usuarioPrevision)
                                        else:
                                            result10={"result":"-100"}
                                        
                                        
                                        #datos asignaciones Colacion, Movilizacion, Familiar
                                        #creamos el esquema de los datos
                                        '''
                                        AsignacionColacion=row[25]
                                        AsignacionMovilizacion=row[26]                                
                                        '''
                                        paso=98
                                        colacionUser=ColacionUserSchema(
                                            sociedad_id=sociedadId,
                                            user_id= newUserId,
                                            movilizacion=AsignacionMovilizacion,
                                            colacion=AsignacionColacion,
                                            familiar=0                                       
                                        )
                                        paso=99
                                        result11=ColacionUsuariosController.create_colacion_user(self,colacionUser,creatorUserId)
                                        
                                        
                                        # insertamos losa datos en la tabla usuarios sociedad
                                        paso=100
                                        usuarioSociedad=UsuariosSociedadSchema(
                                            sociedad_id=sociedadId,
                                            user_id=newUserId
                                        )
                                        paso=101
                                        result12=CargaMasivaController.asignate_user_society(self,usuarioSociedad, creatorUserId)
                                        
                                        #conformamos los estatus de insercion de los datos del archivo
                                        estatusOk="Se insertaron los siguientes datos: Personales"
                                        estatusError=". No se insertaron los siguientes datos"
                                        
                                        if (result2["result"]=="1"):
                                            paso=102
                                            estatusOk = estatusOk + ", Contacto"
                                        else:   
                                            paso=103
                                            estatusError= " Contacto"

                                        if (result3["result"] is not None):
                                            if (result3["result"]=="1"):
                                                paso=104
                                                estatusOk = estatusOk + ", Localizacion"
                                            elif (result3["result"]=="-5"):   
                                                paso=105
                                                estatusError= ", Localizacion La Region y la Comuna están erradas"  
                                            elif (result3["result"]=="-6"):   
                                                paso=105
                                                estatusError= ", Localizacion La Region esta errada y la Comuna esta correcta"  
                                            elif (result3["result"]=="-7"):   
                                                paso=105
                                                estatusError= ", Localizacion La Region esta correcta y la Comuna esta errada"  
                                            elif (result3["result"]=="-8"):   
                                                paso=105
                                                estatusError= ", Localizacion Error no controlado"  

                                        else:
                                            paso=105
                                            estatusError= ", Localizacion Error en el valor de la Region y la Comuna"  
                                                                                        
                                          
                                                                                         
                                        if (result4["result"]=="1"):
                                            paso=106
                                            estatusOk = estatusOk + ", Datos Laborales"
                                        else:   
                                            paso=107
                                            estatusError= ", Datos Laborales"    


                                        if (result5["result"]=="1"):
                                            paso=108
                                            estatusOk = estatusOk + ", Datos de Pago"
                                        else:   
                                            paso=109
                                            estatusError= ", Datos de Pago"         

                                        if (result6 is not None):
                                            paso=110
                                            if (result6["result"]=="1"):
                                                paso=111
                                                estatusOk = estatusOk + ", Campos Adicionales"
                                            else:   
                                                paso=112
                                                estatusError= ", No se procesaron campos adicionales"                                

                                        if (result7 is not None):
                                            paso=113
                                            if (result6["result"]=="1"):
                                                paso=114
                                                estatusOk = estatusOk + ", Usuarios Grupo"
                                            else:   
                                                paso=115
                                                estatusError= ", No se procesaron Usuarios Grupos" 
                                                
                                        if (result8["result"]=="1"): 
                                            paso=116                                       
                                            estatusOk = estatusOk + ", AFC"
                                        else:   
                                            paso=117
                                            estatusError= ", No se procesaron AFC" 
                                            
                                        if (result9 is not None):
                                            paso=118
                                            if (result9["result"]=="1"):   
                                                paso=119                                     
                                                estatusOk = estatusOk + ", AFP"
                                            else:   
                                                paso=120
                                                estatusError= ", No se procesaron AFP"  
                                                                                    
                                        if (result10 is not None):
                                            paso=121
                                            if (result10["result"]=="1"): 
                                                paso=122                                       
                                                estatusOk = estatusOk + ", Prevision Salud"
                                            else:   
                                                paso=123
                                                estatusError= ", No se procesaron datos de Prevision Salud"  

                                        if (result11 is not None):
                                            paso=124                                       
                                            if (result11["result"]=="1"):  
                                                paso=125                                      
                                                estatusOk = estatusOk + ", Datos de Colacion"
                                            else:   
                                                paso=126
                                                estatusError= ", No se procesaron datos de Colacion"                                          
                                                
                                        if (result12 is not None):
                                            paso=127
                                            if (result12["result"]=="1"):  
                                                paso=128                                      
                                                estatusOk = estatusOk + ", Se asocio el empleado a la sociedad"
                                            else:   
                                                paso=129
                                                estatusError= ", No se asocio el empleado a la sociedad"    
                                                                            
                                    elif (result['result']=="-2"):
                                        paso=130
                                        estatusOk=" -2"
                                        estatusError="No se insertaron ninguno de los datos del empleado. Ya existe este username en el sistema"                        

                                    elif (result['result']=="-3"):
                                        paso=131
                                        estatusOk=" -3"
                                        estatusError="No se insertaron ninguno de los datos del empleado. Ya existe este Rut en el sistema"  
                                    elif (result['result']=="-1"):
                                        paso=131
                                        estatusOk="Error "+ result['cadenaError']
                                        estatusError="No se insertaron ninguno de los datos del empleado. Ya existe este Rut en el sistema"  


                                else:
                                    paso=132
                                    estatusOk=" nulo"
                                    estatusError="No se insertaron ninguno de los datos del empleado. Error de validacion de datos " + causa

                                #creamos la data que representa una fila en el archivo de salida
                                if (estatusError == ". No se insertaron los siguientes datos"):
                                    paso=133
                                    estatusError=""

                                
                                dataR={
                                    "rut":rutv,
                                    "nombres":nombresV,
                                    "apellidos":(" ".join([apellidoPaterno, apellidoMaterno])),
                                    "estatus":result['result']+"  "+estatusOk + " " + estatusError
                                }
                                paso=134
                                resultados.append(dataR) 
                                      
                        else:
                            break   

                    paso=200            
                    if (len(resultados)>0):
                        main_file = os.path.abspath(__file__)
                        app_dir = os.path.dirname(main_file)+"/.."

                        #creamos el archivo de salida
                        #archivoSalida = f"/{app_dir}/{ruta}/{uuid.uuid4()}.csv"
                        paso=201
                        archivoSalida = f"{os.getenv('MASSIVE_USERS_OUTPUT')}/{uuid.uuid4()}.csv"
                        # Abrir el archivo en modo de escritura
                        paso=202
                        with open(archivoSalida, 'w', newline='') as csvfile:
                            paso=203
                            # Crear un escritor CSV
                            writer = csv.DictWriter(csvfile, fieldnames=resultados[0].keys())
                            
                            # Escribir la cabecera
                            paso=204
                            writer.writeheader()
                            
                            # Escribir los datos
                            paso=205
                            for row in resultados:
                                paso=206
                                if (row is not None):
                                    paso=207
                                    writer.writerow(row)
                                else:
                                    break
                        paso=208
                        archivoSalidaFinal=archivoSalida.replace("/var/www/html/","/")
                        return ({"result":"1","estado":"archivo_procesado","fileResult":archivoSalidaFinal})            
                                
                    else:
                        return ({"result":"-2","estado":"archivo_no procesado, no se arrojo resultados del procesamiento de los registros"})
                else:
                    #no hay filas que procesar
                    paso=209
                    return ({"result":"-4","estado":"archivo_no procesado, no se arrojo resultados del procesamiento de los registros"})                
            except ValueError as e:
                    # ocurrio un error y devolvemos el estado
                    return( {"result":"-3","cadenaError": f"Error {str(e)} paso {paso}"})  
                
        else:
            return ({"result":"-1","estado":"archivo_no permitido","Archivos Permitidos":list(permitedExtensionMassiveUsers)})        